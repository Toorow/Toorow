"""toorow -- CSV/Excel governed import parser + ledger orchestration (Story 12.9).

Turns a raw CSV or Excel upload into a validated, versioned candidate dataset that
flows through the 12.8 immutable import ledger and the 12.5 atomic publication gates.

Architecture boundaries enforced HARD in this module:
  AD-2  : source-agnostic. The feed format (csv/excel) is resolved once at the
           function boundary and never branched on inside the domain logic. No
           provider name, no connector slug, no import from server/modules/*.
  AD-8  : dbt is the ONLY writer of analytical marts. This module:
             (a) writes NOTHING to a mart; the raw landing is project-scoped
                 ``managed_feed_<datastream_id>`` opened via the 12.8 ledger seam;
             (b) delegates publication to ``core.datastream_publication`` via the
                 12.8 ``open_import`` / ``record_rows`` / ``evaluate_rejection_gate``
                 / ``mark_outcome`` contract -- it never re-implements ledger semantics.
  AD-9  : null is NEVER coerced to 0; rejected rows are honest per-row evidence; a
           zero-row parse is an explicit outcome, not a silent empty list.
  NFR14 : the ledger row + isolated 12.5 candidate exist BEFORE any row is written;
           a parse failure leaves no published mutation.
  NFR15 : replace is the safe default; append is blocked until a stable-key contract
           exists (12.9 scope: Replace only).

The module is structured in two clean layers:
  1. PARSE layer (pure Python, no I/O, unit-testable on in-memory bytes):
       ``detect_format``, ``validate_upload_meta``, ``parse_csv``, ``parse_excel``,
       ``build_preview``, ``validate_import_contract``.
  2. ORCHESTRATE layer (consumes 12.8 public contract, pg-gated):
       ``run_import`` -- drives parse -> open_import -> record_rows -> gate.
       ``version_contract`` -- idempotent insert/dedup into
       ``app.csv_excel_import_contracts`` (AC2) so import configs are reusable.

Per-row rejection (AC3, C2): the parse layer emits ``RejectedRow`` evidence for
rows that fail the contract (date-format mismatch, non-coercible values in a
confirmed-typed column, short/over-long rows). ``run_import`` forwards that evidence
to ``record_rows`` and the UNMOCKED 12.8 rejection gate blocks above the governed
threshold.

The parsing contract is VERSIONED (``csv_excel_import_contracts``) via migration 078
so import configurations are reusable and reviewable (AC2). ``version_contract``
computes a stable ``fingerprint`` and threads ``import_contract_id`` onto the ledger
row so an import is traceable to the exact config version used.

Empty-file / zero-row policy (AC4):
  ``run_import`` returns ``{"blocked": True, "reason": "empty_import"}`` by default.
  An intentional empty publish requires BOTH:
    * ``force_empty_publish=True`` in the call, AND
    * the project preference ``allow_empty_publication=True``.
  Absent both, the ledger is NOT opened and the prior published state is untouched.
  The ``build_preview`` function always reports ``row_count=0`` honestly.

Append unavailability (AC5):
  Calling ``run_import`` with ``write_mode='append'`` raises ``AppendUnavailable``.
  The error carries a ``repair`` explanation that Replace is the safe supported mode.
  No append path exists in this story (deferred to a keyed-dedup contract).

Publication honesty (Phase B):
  A passing rejection gate leaves the ledger at ``written``; ``run_import`` returns
  ``{"outcome": "written_pending_publication", "published": False, ...}``. The physical
  row write (``open_managed_landing``), the full 12.5 DQ gates, and the pointer-swap
  ``commit_publication`` are HONEST Phase B (no live warehouse in dev) -- this module
  NEVER fabricates a publish.
"""

from __future__ import annotations

import csv as _csv
import hashlib
import io
import json
import logging
import re as _re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable

from ulid import ULID

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Closed vocabularies (mirrors 12.8 managed_feed_ledger).
# ---------------------------------------------------------------------------

FORMAT_CSV = "csv"
FORMAT_EXCEL = "excel"
SUPPORTED_FORMATS = frozenset({FORMAT_CSV, FORMAT_EXCEL})

WRITE_MODE_REPLACE = "replace"
WRITE_MODE_APPEND = "append"

# ---------------------------------------------------------------------------
# Size + structural guards (DoS / sanity limits).
# These are documented defaults — overridable per project preference in Phase B.
# ---------------------------------------------------------------------------

MAX_FILE_BYTES = 50 * 1024 * 1024  # 50 MB — generous for real data, DoS-resistant
MAX_ROWS = 500_000  # bounded full-file scan
MAX_COLUMNS = 500  # wide-enough for real datasets

# Excel-specific guards (same spirit as mediaplan_import.py limits).
MAX_SHEET_ROWS = 500_000
MAX_SHEET_COLUMNS = 500

# Bounded preview returned WITHOUT publishing.
PREVIEW_ROW_LIMIT = 100

# Samples collected per column for type inference (cap; only [:10] are stored).
SAMPLE_CAP = 100

# Supported text encodings tried in order (CSV only).
_ENCODINGS = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]

# Allowed delimiters for auto-detection (CSV).
_DELIMITERS = [",", ";", "\t", "|"]

# Date format tokens we recognise in contract declarations.
SUPPORTED_DATE_FORMATS = frozenset({"%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y%m%d"})

# Locale codes we understand for numeric formatting.
SUPPORTED_LOCALES = frozenset({"en", "fr"})

# Per-column types an operator may DECLARE in the contract (`column_types`). A
# declared type is AUTHORITATIVE and overrides the leading-value inference -- it kills
# the false-rejects the heuristic causes on columns whose head merely LOOKS typed (an
# ID/SKU column starting "00123" -> integer; an 8-digit code starting "20260101" ->
# date). 'text' declared on such a column makes it accept everything.
SUPPORTED_COLUMN_TYPES = frozenset({"integer", "decimal", "date", "boolean", "text"})

# ---------------------------------------------------------------------------
# Per-row rejection rule codes (stable; stored on managed_feed_rejected_rows).
# ---------------------------------------------------------------------------

RULE_DATE_FORMAT = "date_format"  # value does not parse under contract date_format
RULE_TYPE_MISMATCH = "type_mismatch"  # non-coercible value in a confirmed-typed column
RULE_SHORT_ROW = "short_row"  # row has fewer fields than the header
RULE_LONG_ROW = "long_row"  # row has more fields than the header (extra data)

# ---------------------------------------------------------------------------
# Typed error hierarchy (stable codes -> callers map to HTTP status).
# ---------------------------------------------------------------------------


class CsvExcelImportError(Exception):
    """Base for all import-layer errors. Carries a stable ``code``."""

    def __init__(self, code: str, detail: str = "", *, repair: dict | None = None) -> None:
        super().__init__(code if not detail else f"{code}: {detail}")
        self.code = code
        self.detail = detail
        self.repair = repair or {}


class UnsupportedFileType(CsvExcelImportError):
    """The upload is neither CSV nor Excel (-> 422)."""

    def __init__(self, detail: str = "") -> None:
        super().__init__(
            "unsupported_file_type",
            detail or "File must be .csv, .tsv, .txt (CSV) or .xlsx, .xls (Excel)",
        )


class FileTooLarge(CsvExcelImportError):
    """The upload exceeds the size guard (-> 422)."""

    def __init__(self, size_bytes: int, max_bytes: int) -> None:
        super().__init__(
            "file_too_large",
            f"File size {size_bytes} bytes exceeds maximum {max_bytes} bytes",
            repair={"compress_or_split": True},
        )


class EncodingError(CsvExcelImportError):
    """The file could not be decoded with any supported encoding (-> 422)."""

    def __init__(self, tried: list[str]) -> None:
        super().__init__(
            "encoding_error",
            f"File could not be decoded. Tried encodings: {tried}. Re-save as UTF-8.",
            repair={"re_encode_as_utf8": True},
        )


class EmptyFile(CsvExcelImportError):
    """The file has no content at all (zero bytes or all whitespace) (-> 422)."""

    def __init__(self) -> None:
        super().__init__(
            "empty_file",
            "The uploaded file is empty (zero bytes or blank).",
            repair={"upload_non_empty_file": True},
        )


class DuplicateColumns(CsvExcelImportError):
    """The header row contains duplicate column names (-> 422)."""

    def __init__(self, duplicates: list[str]) -> None:
        super().__init__(
            "duplicate_columns",
            f"Duplicate column names found: {', '.join(duplicates)}",
            repair={"rename_duplicate_columns": True},
        )


class NoHeaderRow(CsvExcelImportError):
    """The file has content but no valid header row (-> 422)."""

    def __init__(self) -> None:
        super().__init__(
            "no_header_row",
            "Could not detect a header row. Ensure the first row contains column names.",
            repair={"add_header_row": True},
        )


class SheetNotFound(CsvExcelImportError):
    """The requested sheet name does not exist in the workbook (-> 422)."""

    def __init__(self, sheet: str, available: list[str]) -> None:
        super().__init__(
            "sheet_not_found",
            f"Sheet '{sheet}' not found. Available: {available}",
            repair={"pick_available_sheet": available},
        )


class FormulaInCells(CsvExcelImportError):
    """Cells contain formula strings instead of computed values (-> 422).

    The import contract must specify ``formulas_as_values=True`` (data_only mode)
    or the file must be re-exported with values only.
    """

    def __init__(self, count: int) -> None:
        super().__init__(
            "formula_in_cells",
            f"{count} cell(s) contain formula strings (=...). "
            "Re-export the workbook as values-only, or set formulas_as_values=true "
            "in the import contract.",
            repair={"export_values_only": True, "or_set_formulas_as_values": True},
        )


class AppendUnavailable(CsvExcelImportError):
    """Append write mode is requested but no stable key + dedup contract exists.

    Replace is the safe, supported mode for 12.9.
    """

    def __init__(self) -> None:
        super().__init__(
            "append_unavailable",
            "Append mode is unavailable: no stable key and deduplication contract "
            "exists for this datastream. Use Replace (write_mode='replace'), which "
            "validates the full candidate before the active pointer changes.",
            repair={"use_replace_mode": True},
        )


class EmptyImportBlocked(CsvExcelImportError):
    """Zero rows were parsed and publication is blocked by default (-> 422)."""

    def __init__(self) -> None:
        super().__init__(
            "empty_import_blocked",
            "The file produced zero rows after parsing. Publication is blocked by "
            "default. To intentionally publish an empty dataset, call again with "
            "force_empty_publish=True and ensure the project preference "
            "'allow_empty_publication' is enabled.",
            repair={
                "add_data_rows": True,
                "or_set_force_empty_publish_with_project_preference": True,
            },
        )


class InvalidImportContract(CsvExcelImportError):
    """The import contract is malformed or contains unsupported settings (-> 422)."""

    def __init__(self, detail: str) -> None:
        super().__init__("invalid_import_contract", detail)


# ---------------------------------------------------------------------------
# Data structures (pure Python, no I/O).
# ---------------------------------------------------------------------------


@dataclass
class ColumnSpec:
    """One column resolved from parsing."""

    name: str  # header label as-found in the file
    index: int  # 0-based column index
    detected_type: str  # "text" | "integer" | "decimal" | "date" | "boolean" | "mixed"
    date_format: str | None = None  # detected or contract-declared date format
    null_count: int = 0
    sample_values: list[Any] = field(default_factory=list)


@dataclass
class RejectedRow:
    """A row that failed a validation rule."""

    row_number: int  # 1-based, matching the file line number (header = 1)
    field_name: str  # column name or "" for row-level issues
    rule: str  # stable rule code (e.g. "date_format", "type_mismatch")
    reason: str  # human-readable explanation
    rejected_value: str  # bounded/redacted original value (max 500 chars)


@dataclass
class ParseResult:
    """Output of parse_csv / parse_excel: accepted rows + evidence."""

    rows: list[dict[str, Any]]  # accepted rows as column->value dicts
    rejected: list[RejectedRow]  # per-row rejection detail
    columns: list[ColumnSpec]  # resolved column specs
    encoding: str  # encoding used (CSV) or "xlsx"/"xls" (Excel)
    delimiter: str | None  # delimiter used (CSV) or None (Excel)
    sheet_name: str | None  # sheet parsed (Excel) or None (CSV)
    detected_row_count: int  # total data rows seen (accepted + rejected)
    content_hash: str  # SHA-256 of the raw input bytes
    # Story 22.10 (AD-4): an OPTIONAL parse-level metadata block a reshape producer
    # may surface (e.g. the media-plan top ``label: value`` block). Additive with a
    # default so the CSV/Excel parsers (Story 12.9) are unaffected.
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class ImportPreview:
    """Bounded preview returned by build_preview WITHOUT any publish."""

    format: str  # "csv" or "excel"
    encoding: str
    delimiter: str | None
    sheet_name: str | None
    columns: list[ColumnSpec]
    row_count: int  # total accepted rows
    rejected_count: int
    preview_rows: list[dict[str, Any]]  # first PREVIEW_ROW_LIMIT accepted rows
    content_hash: str
    contract_version_id: str | None  # if a contract was applied


# ---------------------------------------------------------------------------
# Pure helper: content hash.
# ---------------------------------------------------------------------------


def _content_hash(data: bytes) -> str:
    """SHA-256 hex of the raw upload bytes (64 hex chars)."""
    return hashlib.sha256(data).hexdigest()


# ---------------------------------------------------------------------------
# Format detection (pure, no I/O).
# ---------------------------------------------------------------------------

_CSV_EXTENSIONS = frozenset({".csv", ".tsv", ".txt"})
_EXCEL_EXTENSIONS = frozenset({".xlsx", ".xls"})

# Excel magic bytes (first 4 bytes of the file).
_XLSX_MAGIC = b"PK\x03\x04"  # ZIP-based (xlsx, xlsm, ...)
_XLS_MAGIC = b"\xd0\xcf\x11\xe0"  # OLE2 compound document (legacy xls)


def detect_format(
    filename: str | None,
    data: bytes,
) -> str:
    """Return ``FORMAT_CSV`` or ``FORMAT_EXCEL``; raise ``UnsupportedFileType`` otherwise.

    Detection order:
      1. Magic bytes (content wins over filename when unambiguous).
      2. File extension (lower-cased).
    """
    if data[:4] == _XLSX_MAGIC or data[:4] == _XLS_MAGIC:
        return FORMAT_EXCEL
    ext = ""
    if filename:
        name = filename.strip().lower()
        for suffix in _EXCEL_EXTENSIONS:
            if name.endswith(suffix):
                return FORMAT_EXCEL
        for suffix in _CSV_EXTENSIONS:
            if name.endswith(suffix):
                return FORMAT_CSV
        dot = name.rfind(".")
        if dot >= 0:
            ext = name[dot:]
    raise UnsupportedFileType(f"Could not determine file type from content or extension '{ext}'")


# ---------------------------------------------------------------------------
# Upload meta validation (pure).
# ---------------------------------------------------------------------------


def validate_upload_meta(
    data: bytes,
    *,
    max_bytes: int = MAX_FILE_BYTES,
) -> None:
    """Validate raw upload bytes before parsing.

    Raises:
      EmptyFile   -- zero bytes or all whitespace.
      FileTooLarge -- size exceeds ``max_bytes``.
    """
    if not data or not data.strip():
        raise EmptyFile()
    if len(data) > max_bytes:
        raise FileTooLarge(len(data), max_bytes)


# ---------------------------------------------------------------------------
# CSV parsing (pure).
# ---------------------------------------------------------------------------


def _detect_encoding(data: bytes) -> tuple[str, str]:
    """Try each encoding in order; return (text, encoding) on first success.

    Raises ``EncodingError`` if all fail.
    """
    for enc in _ENCODINGS:
        try:
            return data.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue
    raise EncodingError(_ENCODINGS)


def _detect_delimiter(sample: str) -> str:
    """Sniff the CSV delimiter from a sample of text (first 8 KB).

    Falls back to comma when ``csv.Sniffer`` cannot decide.
    """
    try:
        dialect = _csv.Sniffer().sniff(sample[:8192], delimiters="".join(_DELIMITERS))
        return dialect.delimiter
    except _csv.Error:
        return ","


# A comma-grouped integer looks like 1,234 or 12,345,678: groups of exactly 3
# digits after the first 1-3 digits. A stray "1,5" is NOT thousands-grouped (the
# group after the comma is not 3 digits) so it stays a decimal candidate (LOW fix:
# integer inference must not strip a decimal comma into a fake integer).
_THOUSANDS_GROUPED = _re.compile(r"^-?\d{1,3}(,\d{3})+$")


def _strip_ws(s: str) -> str:
    """Strip surrounding whitespace + interior thousands spaces (regular + NBSP + thin)."""
    return s.strip().replace(" ", "").replace(" ", "").replace(" ", "")


def _looks_like_thousands_grouped(s: str) -> bool:
    return bool(_THOUSANDS_GROUPED.match(s))


def _coerce_integer(raw: str) -> int:
    """Coerce a stripped string to int (thousands-grouping tolerated). Raises ValueError."""
    s = _strip_ws(raw)
    if _looks_like_thousands_grouped(s):
        s = s.replace(",", "")
    return int(s)


def _coerce_decimal(raw: str) -> float:
    """Coerce a stripped string to float (FR/EN decimal conventions). Raises ValueError."""
    s = _strip_ws(raw)
    if "," in s and "." in s:
        # Both present: the LAST separator is the decimal point.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")  # FR: 1.234,56
        else:
            s = s.replace(",", "")  # EN: 1,234.56
    elif "," in s:
        if _looks_like_thousands_grouped(s):
            s = s.replace(",", "")  # 1,234 (thousands)
        else:
            s = s.replace(",", ".")  # 1,5 (FR decimal)
    return float(s)


def _is_coercible(raw: str, detected_type: str, date_format: str | None) -> bool:
    """Return True if ``raw`` is coercible to the column's confirmed type (C2 evidence).

    ``text`` / ``mixed`` never reject (any string is valid text). An empty value is
    honest null, never a rejection (AD-9). ``boolean`` / ``integer`` / ``decimal`` /
    ``date`` reject non-coercible values.
    """
    s = raw.strip()
    if s == "":
        return True
    if detected_type in ("text", "mixed"):
        return True
    if detected_type == "boolean":
        return s.lower() in {"true", "false", "yes", "no", "1", "0", "oui", "non"}
    if detected_type == "integer":
        try:
            _coerce_integer(s)
            return True
        except ValueError:
            return False
    if detected_type == "decimal":
        try:
            _coerce_decimal(s)
            return True
        except ValueError:
            return False
    if detected_type == "date":
        fmts = [date_format] if date_format else sorted(SUPPORTED_DATE_FORMATS)
        for fmt in fmts:
            if not fmt:
                continue
            try:
                datetime.strptime(s, fmt)
                return True
            except ValueError:
                continue
        return False
    return True


def _resolve_column_type(
    col_name: str,
    values: list[Any],
    column_types: dict[str, str] | None,
) -> tuple[str, str | None]:
    """Resolve a column's type. A DECLARED contract type (``column_types[col_name]``)
    is authoritative and short-circuits the leading-value inference -- this is the fix
    for the heuristic's false-rejects on ID/SKU/YYYYMMDD-looking columns. Returns
    ``(type, date_format_or_None)``; a declared column carries no inferred date_format,
    so the caller's global contract ``date_format`` (or the supported-format set)
    governs a declared 'date' column. Undeclared columns fall back to inference.
    """
    declared = (column_types or {}).get(col_name)
    if declared:
        return declared, None
    return _infer_column_type(values)


def _infer_column_type(values: list[Any]) -> tuple[str, str | None]:
    """Establish a column's type from its LEADING non-null value(s).

    Returns ``(detected_type, date_format_or_None)``.

    Design (C2): a full-column "all values must agree" inference cannot detect a
    dirty column -- 2 good integers followed by 4 garbage values would infer ``mixed``
    and reject NOTHING, defeating the per-row rejection contract. Instead the type is
    ESTABLISHED by the leading data (the representative head of the column) and every
    row is then validated against it in ``_validate_rows``; non-conforming rows become
    honest ``RejectedRow`` evidence that drives the 12.8 rejection-threshold gate.

    The type is taken from the FIRST non-null value, tried in priority order
    date > integer > decimal > boolean > text (integer before boolean so a leading
    ``1`` is a number, not a flag; the CONTRACT ``date_format`` override still forces a
    strict date format downstream). A free-text head yields ``text`` -- which never
    rejects -- so a genuine notes column produces no false rejections.
    """
    non_null = [str(v).strip() for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "text", None

    head = non_null[0]

    # Date first (a value like 2026-01-01 must not be read as an integer/decimal).
    for fmt in sorted(SUPPORTED_DATE_FORMATS):
        try:
            datetime.strptime(head, fmt)
            return "date", fmt
        except ValueError:
            pass
    # Integer before boolean: a leading "1"/"0" is a number, not a flag.
    try:
        _coerce_integer(head)
        return "integer", None
    except ValueError:
        pass
    try:
        _coerce_decimal(head)
        return "decimal", None
    except ValueError:
        pass
    if head.lower() in {"true", "false", "yes", "no", "oui", "non"}:
        return "boolean", None
    return "text", None


def _check_formula_cells(
    text: str,
    *,
    delimiter: str,
    max_scan: int = 10_000,
) -> int:
    """Count fields that look like formula strings (start with '=').

    H1 fix: split each scanned line on the ALREADY-SNIFFED ``delimiter`` (not a
    hardcoded tab) so a comma/semicolon/pipe CSV exported with formulas (e.g.
    ``name,=SUM(B1)``) is caught. Scans up to ``max_scan`` lines; one formula in the
    first 10 000 lines is enough evidence to block and ask for a values-only export.
    """
    count = 0
    for i, line in enumerate(text.splitlines()):
        if i >= max_scan:
            break
        for token in line.split(delimiter):
            if token.strip().startswith("="):
                count += 1
    return count


def _strip_bom(value: str) -> str:
    """Drop a leading UTF-8/UTF-16 BOM char from a header/cell label.

    Defends against DOUBLE-BOM exports (a literal ``﻿`` in the source that then
    gets ``utf-8-sig`` re-encoded): the ``utf-8-sig`` decode strips only the FIRST BOM,
    so the first header would otherwise read ``﻿col_a``. This keeps column names
    honest without changing the detected encoding.
    """
    return value.lstrip("﻿")


def _truncate_value(v: Any, max_len: int = 500) -> str:
    """Bounded/redacted string for rejected-row storage (AD-9)."""
    s = str(v) if v is not None else ""
    if len(s) > max_len:
        return s[:max_len] + "…"
    return s


def _validate_rows(
    accepted_rows: list[tuple[int, list[str]]],
    col_specs: list[ColumnSpec],
    col_names: list[str],
) -> tuple[list[dict[str, Any]], list[RejectedRow]]:
    """Second pass: per-row validation against the inferred column specs (C2).

    ``accepted_rows`` is a list of ``(line_number, cells)`` -- the raw non-blank rows
    with their 1-based file line number. Returns ``(accepted, rejected)``:
      * a row with MORE cells than columns -> ``long_row`` (extra data would be lost);
      * a value non-coercible to a column's confirmed (non-text) type -> the row is
        rejected on the FIRST offending field (``type_mismatch`` / ``date_format``).
    A row with FEWER cells than the header is NOT rejected outright (short rows are
    common padding); missing trailing fields become honest nulls. Only genuinely
    truncated rows (fewer cells AND a non-empty value expected) are not distinguishable
    here, so short rows pad with null (AD-9) -- over-long rows ARE rejected because
    they signal a delimiter/structure error.
    """
    accepted: list[dict[str, Any]] = []
    rejected: list[RejectedRow] = []
    n_cols = len(col_names)
    for line_no, cells in accepted_rows:
        # Over-long row: more cells than columns -> structural error, reject.
        if len(cells) > n_cols:
            rejected.append(
                RejectedRow(
                    row_number=line_no,
                    field_name="",
                    rule=RULE_LONG_ROW,
                    reason=(
                        f"row has {len(cells)} fields but the header declares {n_cols} "
                        "columns (delimiter or structure error)"
                    ),
                    rejected_value=_truncate_value(cells),
                )
            )
            continue

        row_dict: dict[str, Any] = {}
        row_rejected = False
        for i, col_name in enumerate(col_names):
            raw_val = cells[i].strip() if i < len(cells) else ""
            spec = col_specs[i]
            if not _is_coercible(raw_val, spec.detected_type, spec.date_format):
                rule = (
                    RULE_DATE_FORMAT
                    if spec.detected_type == "date"
                    else RULE_TYPE_MISMATCH
                )
                reason = (
                    f"value does not match date format "
                    f"{spec.date_format!r} for column '{col_name}'"
                    if spec.detected_type == "date"
                    else (
                        f"value is not a valid {spec.detected_type} for column "
                        f"'{col_name}'"
                    )
                )
                rejected.append(
                    RejectedRow(
                        row_number=line_no,
                        field_name=col_name,
                        rule=rule,
                        reason=reason,
                        rejected_value=_truncate_value(raw_val),
                    )
                )
                row_rejected = True
                break
            row_dict[col_name] = raw_val if raw_val != "" else None
        if not row_rejected:
            accepted.append(row_dict)
    return accepted, rejected


def parse_csv(
    data: bytes,
    *,
    delimiter: str | None = None,
    date_format: str | None = None,
    locale: str | None = None,
    header_row: int = 1,
    max_rows: int = MAX_ROWS,
    max_columns: int = MAX_COLUMNS,
    check_formulas: bool = True,
    column_types: dict[str, str] | None = None,
) -> ParseResult:
    """Parse CSV bytes into a ``ParseResult`` with per-row rejection evidence.

    Args:
      data         : raw upload bytes.
      delimiter    : override auto-detection (None = sniff).
      date_format  : override date-format detection for the whole file.
      locale       : "en" or "fr" (numeric decimal convention).
      header_row   : 1-based row number of the column headers (default: 1).
      max_rows     : row limit (DoS guard).
      max_columns  : column limit (DoS guard).
      check_formulas : block on formula strings.

    Raises:
      EncodingError, NoHeaderRow, DuplicateColumns, FormulaInCells, too_many_columns.
    """
    ch = _content_hash(data)
    text, enc = _detect_encoding(data)

    # H1: sniff the delimiter FIRST, then run the formula check on the parsed fields.
    used_delim = delimiter or _detect_delimiter(text)

    # Formula check (CSV/TSV exported from a spreadsheet with formulas). Runs on the
    # already-sniffed delimiter so a comma/semicolon/pipe CSV is not missed.
    if check_formulas:
        formula_count = _check_formula_cells(text, delimiter=used_delim)
        if formula_count > 0:
            raise FormulaInCells(formula_count)

    reader = _csv.reader(io.StringIO(text), delimiter=used_delim)
    rows_raw: list[list[str]] = list(reader)
    if not rows_raw:
        return ParseResult(
            rows=[],
            rejected=[],
            columns=[],
            encoding=enc,
            delimiter=used_delim,
            sheet_name=None,
            detected_row_count=0,
            content_hash=ch,
        )

    if header_row < 1 or header_row > len(rows_raw):
        raise NoHeaderRow()

    header = rows_raw[header_row - 1]
    if not header or all(h.strip() == "" for h in header):
        raise NoHeaderRow()

    # Normalise header labels: strip whitespace + any residual BOM (double-BOM export).
    col_names = [_strip_bom(h.strip()) for h in header]

    # LOW: duplicate-column check BEFORE the column-count DoS guard so a maliciously
    # wide duplicate header is still reported as a duplicate (the more actionable error).
    seen: dict[str, int] = {}
    for name in col_names:
        key = name.casefold()
        seen[key] = seen.get(key, 0) + 1
    duplicates = [name for name, count in seen.items() if count > 1]
    if duplicates:
        raise DuplicateColumns(duplicates)

    if len(col_names) > max_columns:
        raise CsvExcelImportError(
            "too_many_columns",
            f"File has {len(col_names)} columns; maximum is {max_columns}.",
        )

    # First pass: collect samples for type inference.
    samples: dict[int, list[Any]] = {i: [] for i in range(len(col_names))}
    non_blank: list[tuple[int, list[str]]] = []  # (1-based line number, cells)
    data_rows = rows_raw[header_row:]
    for raw_row_idx, raw_row in enumerate(data_rows):
        if raw_row_idx >= max_rows:
            logger.warning("csv parse: row limit %d reached, truncating", max_rows)
            break
        if all(cell.strip() == "" for cell in raw_row):
            continue
        line_no = header_row + raw_row_idx + 1  # 1-based file line
        # Defensive trailing-empty trim: a ragged trailing delimiter should not be
        # mistaken for an over-long row. A trailing cell with real data is preserved,
        # so a genuinely over-long row still rejects (long_row).
        cells = list(raw_row)
        while len(cells) > len(col_names) and cells[-1].strip() == "":
            cells.pop()
        non_blank.append((line_no, cells))
        for i in range(len(col_names)):
            raw_val = cells[i].strip() if i < len(cells) else ""
            if raw_val != "" and len(samples[i]) < SAMPLE_CAP:
                samples[i].append(raw_val)

    col_specs: list[ColumnSpec] = []
    for i, col_name in enumerate(col_names):
        det_type, det_fmt = _resolve_column_type(col_name, samples[i], column_types)
        col_specs.append(
            ColumnSpec(
                name=col_name,
                index=i,
                detected_type=det_type,
                date_format=date_format or det_fmt,
                null_count=0,  # filled after the validation pass
                sample_values=samples[i][:10],
            )
        )

    # Second pass: per-row validation against the resolved specs (C2 evidence).
    accepted, rejected = _validate_rows(non_blank, col_specs, col_names)

    # Finalise null counts on the accepted rows (honest, post-validation).
    for spec in col_specs:
        spec.null_count = sum(1 for r in accepted if r.get(spec.name) is None)

    return ParseResult(
        rows=accepted,
        rejected=rejected,
        columns=col_specs,
        encoding=enc,
        delimiter=used_delim,
        sheet_name=None,
        detected_row_count=len(non_blank),
        content_hash=ch,
    )


# ---------------------------------------------------------------------------
# Excel parsing (pure, uses openpyxl).
# ---------------------------------------------------------------------------


def _load_workbook_safe(data: bytes, *, data_only: bool = True) -> Any:
    """Load an openpyxl workbook from bytes; raise UnsupportedFileType on failure."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:  # pragma: no cover
        raise CsvExcelImportError(
            "openpyxl_unavailable",
            "openpyxl is required for Excel import. Install it: pip install openpyxl",
        ) from exc
    try:
        return openpyxl.load_workbook(io.BytesIO(data), data_only=data_only, read_only=True)
    except Exception as exc:
        raise UnsupportedFileType(f"Could not open as Excel workbook: {exc}") from exc


def _check_excel_formulas(ws: Any, *, max_cells: int = 5000) -> int:
    """Count formula-string cells in the worksheet.

    A cached-formula string (a cell whose value is a ``=...`` string) is caught here
    REGARDLESS of the ``formulas_as_values`` mode: even in ``data_only=True`` mode a
    workbook saved without a compute pass can still surface the formula string, and
    silently importing ``=SUM(B1)`` as a text value is wrong (H1). We scan and report.
    """
    count = 0
    scanned = 0
    for row in ws.iter_rows():
        for cell in row:
            scanned += 1
            if scanned > max_cells:
                return count
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count


def parse_excel(
    data: bytes,
    *,
    sheet_name: str | None = None,
    header_row: int = 1,
    date_format: str | None = None,
    locale: str | None = None,
    formulas_as_values: bool = True,
    max_rows: int = MAX_SHEET_ROWS,
    max_columns: int = MAX_SHEET_COLUMNS,
    column_types: dict[str, str] | None = None,
) -> ParseResult:
    """Parse Excel bytes (.xlsx / .xls) into a ``ParseResult`` with rejection evidence.

    Args:
      data            : raw upload bytes.
      sheet_name      : sheet to read (None = first sheet).
      header_row      : 1-based row number of column headers (default: 1).
      date_format     : override date-format detection.
      locale          : "en" or "fr".
      formulas_as_values : if True (default), load with data_only=True so computed
                         values are returned. A cached formula STRING is always scanned
                         (see ``_check_excel_formulas``) and blocks regardless of mode.
      max_rows/max_columns : DoS guards.

    Raises:
      UnsupportedFileType, SheetNotFound, NoHeaderRow, DuplicateColumns, FormulaInCells.
    """
    ch = _content_hash(data)
    wb = _load_workbook_safe(data, data_only=formulas_as_values)

    # Sheet resolution.
    sheet_names = wb.sheetnames
    if sheet_name is None:
        ws = wb.active
        used_sheet = ws.title if ws is not None else (sheet_names[0] if sheet_names else "Sheet1")
    else:
        if sheet_name not in sheet_names:
            raise SheetNotFound(sheet_name, sheet_names)
        ws = wb[sheet_name]
        used_sheet = sheet_name

    if ws is None:
        raise CsvExcelImportError("no_active_sheet", "Workbook has no active sheet.")

    # H1: scan for cached formula STRINGS regardless of formulas_as_values. A cell
    # whose value is "=..." must never be imported verbatim as text.
    formula_count = _check_excel_formulas(ws)
    if formula_count > 0:
        raise FormulaInCells(formula_count)

    # Read rows into memory (bounded by max_rows guard). The +header_row headroom lets
    # us reach header_row data rows after the header offset; the second break below is
    # the authoritative data-row cap (LOW: single guard, not a double one that
    # miscounts when header_row > 1).
    all_rows: list[list[Any]] = []
    row_cap = max_rows + header_row
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= row_cap:
            logger.warning("excel parse: row limit %d reached, truncating", max_rows)
            break
        all_rows.append(list(row))

    if len(all_rows) < header_row:
        raise NoHeaderRow()

    # Header resolution.
    header_cells = all_rows[header_row - 1]
    col_names_raw = [
        _strip_bom(str(c).strip()) if c is not None else "" for c in header_cells
    ]
    # Drop trailing empty columns.
    while col_names_raw and col_names_raw[-1] == "":
        col_names_raw.pop()
    if not col_names_raw:
        raise NoHeaderRow()

    # Duplicate check BEFORE the column-count DoS guard (LOW: report the actionable
    # duplicate error first).
    seen_c: dict[str, int] = {}
    for h in col_names_raw:
        key = h.casefold()
        seen_c[key] = seen_c.get(key, 0) + 1
    dups = [h for h, cnt in seen_c.items() if cnt > 1]
    if dups:
        raise DuplicateColumns(dups)

    if max_columns and len(col_names_raw) > max_columns:
        raise CsvExcelImportError(
            "too_many_columns",
            f"Sheet has {len(col_names_raw)} columns; maximum is {max_columns}.",
        )

    col_count = len(col_names_raw)
    samples: dict[int, list[Any]] = {i: [] for i in range(col_count)}
    non_blank: list[tuple[int, list[str]]] = []
    data_rows = all_rows[header_row:]
    for raw_row_idx, raw_row in enumerate(data_rows):
        if raw_row_idx >= max_rows:
            break
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in raw_row):
            continue
        line_no = header_row + raw_row_idx + 1
        # Stringify cells so the shared validation/coercion path works for Excel too.
        cells = ["" if c is None else str(c) for c in raw_row]
        # Drop trailing EMPTY cells: openpyxl pads every data row to the sheet's max
        # column extent, so a row that only has blank padding beyond the (trimmed)
        # header must NOT be mistaken for an over-long row. A trailing cell with real
        # data is preserved, so a genuinely over-long row still rejects (long_row).
        while len(cells) > col_count and cells[-1].strip() == "":
            cells.pop()
        non_blank.append((line_no, cells))
        for i in range(col_count):
            raw_val = cells[i].strip() if i < len(cells) else ""
            if raw_val != "" and len(samples[i]) < SAMPLE_CAP:
                samples[i].append(raw_val)

    col_specs: list[ColumnSpec] = []
    for i, col_name in enumerate(col_names_raw):
        det_type, det_fmt = _resolve_column_type(col_name, samples[i], column_types)
        col_specs.append(
            ColumnSpec(
                name=col_name,
                index=i,
                detected_type=det_type,
                date_format=date_format or det_fmt,
                null_count=0,
                sample_values=samples[i][:10],
            )
        )

    accepted, rejected = _validate_rows(non_blank, col_specs, col_names_raw)
    for spec in col_specs:
        spec.null_count = sum(1 for r in accepted if r.get(spec.name) is None)

    return ParseResult(
        rows=accepted,
        rejected=rejected,
        columns=col_specs,
        encoding="xlsx",
        delimiter=None,
        sheet_name=used_sheet,
        detected_row_count=len(non_blank),
        content_hash=ch,
    )


# ---------------------------------------------------------------------------
# Import contract validation (pure).
# ---------------------------------------------------------------------------

# The import contract is a versioned JSON document that governs how a recurring
# CSV/Excel upload is parsed.  It is STORED in ``app.csv_excel_import_contracts``
# (migration 078) so the same configuration is reusable and auditable across runs.
#
# Minimal contract shape:
#   {
#     "format": "csv" | "excel",
#     "delimiter": "," | ";" | "\t" | "|" | null,  -- CSV only; null = auto-detect
#     "encoding": "utf-8" | ...,                    -- CSV only; null = auto-detect
#     "sheet_name": "Sheet1" | null,                -- Excel only
#     "header_row": 1,                              -- 1-based (default 1)
#     "date_format": "%Y-%m-%d" | null,             -- null = auto-detect
#     "locale": "en" | "fr" | null,
#     "formulas_as_values": true,                   -- Excel only (default true)
#     "write_mode": "replace",                      -- only "replace" in 12.9
#     "column_types": {                             -- optional; declared type per column
#         "id": "text", "spend": "decimal", "day": "date"   -- overrides inference
#     }
#   }


def validate_import_contract(contract: dict[str, Any]) -> dict[str, Any]:
    """Validate and normalise a parsing/import contract dict.

    Returns a normalised copy on success.
    Raises ``InvalidImportContract`` (-> 422) on structural / unsupported-value errors.
    """
    if not isinstance(contract, dict):
        raise InvalidImportContract("Contract must be a JSON object.")

    # format.
    fmt = contract.get("format")
    if fmt not in SUPPORTED_FORMATS:
        raise InvalidImportContract(
            f"'format' must be one of {sorted(SUPPORTED_FORMATS)}; got {fmt!r}."
        )

    # write_mode (only replace in 12.9).
    wm = contract.get("write_mode", WRITE_MODE_REPLACE)
    if wm == WRITE_MODE_APPEND:
        raise AppendUnavailable()
    if wm != WRITE_MODE_REPLACE:
        raise InvalidImportContract(
            f"'write_mode' must be 'replace'; got {wm!r}. "
            "Append mode is not available (no stable key contract)."
        )

    # header_row.
    hr = contract.get("header_row", 1)
    if not isinstance(hr, int) or isinstance(hr, bool) or hr < 1:
        raise InvalidImportContract("'header_row' must be a positive integer >= 1.")

    # delimiter (CSV only; None = auto).
    delim = contract.get("delimiter", None)
    if delim is not None and delim not in _DELIMITERS:
        raise InvalidImportContract(
            f"'delimiter' must be one of {_DELIMITERS} or null; got {delim!r}."
        )

    # date_format (optional; None = auto).
    df = contract.get("date_format", None)
    if df is not None and df not in SUPPORTED_DATE_FORMATS:
        raise InvalidImportContract(
            f"'date_format' {df!r} is not in the supported set: {sorted(SUPPORTED_DATE_FORMATS)}."
        )

    # locale (optional).
    loc = contract.get("locale", None)
    if loc is not None and loc not in SUPPORTED_LOCALES:
        raise InvalidImportContract(
            f"'locale' must be one of {sorted(SUPPORTED_LOCALES)} or null; got {loc!r}."
        )

    # sheet_name (Excel only; string or None).
    sn = contract.get("sheet_name", None)
    if sn is not None and not isinstance(sn, str):
        raise InvalidImportContract("'sheet_name' must be a string or null.")

    # formulas_as_values (Excel only; bool, default true).
    fav = contract.get("formulas_as_values", True)
    if not isinstance(fav, bool):
        raise InvalidImportContract("'formulas_as_values' must be a boolean.")

    # column_types (optional): a DECLARED per-column type map {name: type}. A declared
    # type is authoritative over the leading-value inference (kills the ID/YYYYMMDD
    # false-rejects). Sorted into a plain dict so the contract fingerprint is stable.
    ct = contract.get("column_types", None)
    column_types: dict[str, str] = {}
    if ct is not None:
        if not isinstance(ct, dict):
            raise InvalidImportContract(
                "'column_types' must be an object mapping column name -> type, or null."
            )
        for cname, ctype in ct.items():
            if not isinstance(cname, str) or not cname.strip():
                raise InvalidImportContract(
                    "'column_types' keys must be non-empty column names."
                )
            if ctype not in SUPPORTED_COLUMN_TYPES:
                raise InvalidImportContract(
                    f"'column_types[{cname}]' must be one of "
                    f"{sorted(SUPPORTED_COLUMN_TYPES)}; got {ctype!r}."
                )
            column_types[cname] = ctype

    # Return normalised copy.
    return {
        "format": fmt,
        "write_mode": WRITE_MODE_REPLACE,
        "header_row": hr,
        "delimiter": delim,
        "encoding": contract.get("encoding", None),
        "date_format": df,
        "locale": loc,
        "sheet_name": sn,
        "formulas_as_values": fav,
        "column_types": column_types,
    }


def contract_fingerprint(normalised_contract: dict[str, Any]) -> str:
    """Stable SHA-256 fingerprint over the NORMALISED parsing contract (AC2).

    Keys are sorted so the fingerprint is invariant to dict ordering. The same config
    -> the same fingerprint -> de-duplication against the DB UNIQUE index on
    ``(datastream_id, project_id, fingerprint)``.
    """
    canonical = json.dumps(normalised_contract, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Preview builder (pure, no publish).
# ---------------------------------------------------------------------------


def build_preview(
    data: bytes,
    *,
    filename: str | None = None,
    contract: dict[str, Any] | None = None,
    format_hint: str | None = None,
) -> ImportPreview:
    """Produce a bounded preview from upload bytes WITHOUT publishing or writing data.

    This is the first half of the 2-step import flow:
      1. ``build_preview`` -- validate + parse + return bounded preview (no publish).
      2. ``run_import``    -- confirm + open_import + record_rows + publication gate.

    The contract (if provided) is validated first; contract settings override
    auto-detection (delimiter, encoding, sheet_name, header_row, date_format).

    Raises: UnsupportedFileType, EmptyFile, FileTooLarge, EncodingError,
            NoHeaderRow, DuplicateColumns, FormulaInCells, InvalidImportContract.
    """
    validate_upload_meta(data)

    # Resolve format.
    if format_hint and format_hint in SUPPORTED_FORMATS:
        fmt = format_hint
    else:
        fmt = detect_format(filename, data)

    # Validate and normalise the contract if provided.
    norm_contract: dict[str, Any] = {}
    if contract:
        norm_contract = validate_import_contract(contract)
        if norm_contract.get("format") and norm_contract["format"] != fmt:
            raise InvalidImportContract(
                f"Contract format '{norm_contract['format']}' does not match "
                f"detected file format '{fmt}'."
            )

    # Parse.
    if fmt == FORMAT_CSV:
        result = parse_csv(
            data,
            delimiter=norm_contract.get("delimiter"),
            date_format=norm_contract.get("date_format"),
            locale=norm_contract.get("locale"),
            header_row=norm_contract.get("header_row", 1),
            column_types=norm_contract.get("column_types"),
        )
    else:
        result = parse_excel(
            data,
            sheet_name=norm_contract.get("sheet_name"),
            header_row=norm_contract.get("header_row", 1),
            date_format=norm_contract.get("date_format"),
            locale=norm_contract.get("locale"),
            formulas_as_values=norm_contract.get("formulas_as_values", True),
            column_types=norm_contract.get("column_types"),
        )

    return ImportPreview(
        format=fmt,
        encoding=result.encoding,
        delimiter=result.delimiter,
        sheet_name=result.sheet_name,
        columns=result.columns,
        row_count=len(result.rows),
        rejected_count=len(result.rejected),
        preview_rows=result.rows[:PREVIEW_ROW_LIMIT],
        content_hash=result.content_hash,
        contract_version_id=None,  # preview does not version the contract (no publish)
    )


# ---------------------------------------------------------------------------
# Contract versioning (AC2) -- idempotent insert/dedup into the 078 table.
# ---------------------------------------------------------------------------


def _mint_contract_id() -> str:
    return f"cic_{ULID()}"


def version_contract(
    contract: dict[str, Any],
    *,
    datastream_id: str,
    project_id: str,
    actor: str,
    conn,
    label: str | None = None,
) -> str:
    """Persist a versioned parsing contract; return its ``cic_<ULID>`` id (AC2).

    Computes a stable ``fingerprint`` over the NORMALISED contract and performs an
    idempotent insert:
      * a NEW config (unseen fingerprint for this datastream+project) -> a new
        ``cic_<ULID>`` row;
      * an IDENTICAL config (same fingerprint) -> the EXISTING row id is returned
        (``ON CONFLICT (datastream_id, project_id, fingerprint) DO NOTHING`` + re-read),
        so an unchanged contract is never re-versioned.

    The caller owns the transaction (does NOT commit here) so the contract row and the
    ledger row that references it commit together. Immutability of the identity fields
    is enforced by the 078 trigger.
    """
    norm = validate_import_contract(contract)
    fingerprint = contract_fingerprint(norm)
    fmt = norm["format"]
    write_mode = norm["write_mode"]

    with conn.cursor() as cur:
        # Fast path: an identical contract already exists -> return it (dedup).
        cur.execute(
            """
            SELECT id FROM app.csv_excel_import_contracts
            WHERE datastream_id = %s AND project_id = %s AND fingerprint = %s
            """,
            (datastream_id, project_id, fingerprint),
        )
        row = cur.fetchone()
        if row is not None:
            return row[0]

        contract_id = _mint_contract_id()
        cur.execute(
            """
            INSERT INTO app.csv_excel_import_contracts
                (id, datastream_id, project_id, fingerprint, format, write_mode,
                 contract, created_by, label)
            VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s)
            ON CONFLICT (datastream_id, project_id, fingerprint) DO NOTHING
            """,
            (
                contract_id,
                datastream_id,
                project_id,
                fingerprint,
                fmt,
                write_mode,
                json.dumps(norm),
                actor,
                label,
            ),
        )
        # Re-read: our insert may have been a no-op if a concurrent writer won the
        # fingerprint race -- return whichever id is now persisted (dedup guarantee).
        cur.execute(
            """
            SELECT id FROM app.csv_excel_import_contracts
            WHERE datastream_id = %s AND project_id = %s AND fingerprint = %s
            """,
            (datastream_id, project_id, fingerprint),
        )
        row = cur.fetchone()
        if row is None:  # pragma: no cover - the insert above just ran.
            raise CsvExcelImportError(
                "contract_version_failed",
                "contract row not found after insert (unexpected)",
            )
        return row[0]


# ---------------------------------------------------------------------------
# Orchestration layer (consumes 12.8 public contract; pg-gated in production).
# ---------------------------------------------------------------------------

# Success signal on a passing gate: the ledger is 'written' but NOT published. The
# physical row write + 12.5 DQ gates + pointer-swap commit_publication are Phase B.
OUTCOME_WRITTEN_PENDING_PUBLICATION = "written_pending_publication"


def run_import(
    data: bytes,
    *,
    datastream_id: str,
    project_id: str,
    plan_version_id: str,
    mapping_version_id: str,
    projection_plan: dict[str, Any],
    actor: str,
    idempotency_key: str,
    source_metadata: dict[str, Any],
    contract: dict[str, Any],
    conn,  # psycopg connection; caller-managed
    raw_schema: str | None = None,
    write_mode: str = WRITE_MODE_REPLACE,
    force_empty_publish: bool = False,
    preferences: dict[str, Any] | None = None,
    producer: Callable[[bytes], "ParseResult"] | None = None,
) -> dict[str, Any]:
    """Drive a CSV/Excel upload through the 12.8 ledger + 12.5 DQ gates.

    File-source producer path (Story 22.12, AD-1): when ``producer`` is supplied,
    the PARSE STEP delegates to it (a ``file_source_producer.produce`` closure
    returning a ``ParseResult`` with rows keyed by canonical field ids) instead of
    ``parse_csv``/``parse_excel``. Everything downstream -- the 12.8 ledger
    (``open_import``/``record_rows``), the rejection gate, and the landing
    relation -- is REUSED UNCHANGED; no second ledger or landing path is
    introduced. The file-source template is the versioned contract (Story 22.11),
    so this path threads ``source_metadata['import_contract_id']`` (optional)
    rather than versioning a CSV/Excel contract.

    This is the ORCHESTRATION step (the second half of the import flow). The caller
    must have already called ``build_preview`` to confirm the file is valid and the
    operator has reviewed the column mapping.

    The parsing contract is VERSIONED (AC2) via ``version_contract`` and its
    ``cic_<ULID>`` id is threaded onto the ledger row (``import_contract_id``).

    Raises:
      AppendUnavailable         -- write_mode='append' is not supported.
      CsvExcelImportError       -- parse failure (encoding, format, structural).
      ImportPayloadConflict     -- same idempotency key, different payload (-> 409).
      ImportInProgress          -- another import is already active (-> 409).
      InvalidImportContract     -- contract format does not match the actual bytes.

    Returns a dict with, on the passing-gate happy path::

      {
        "ledger": {...},        # the mfl_<ULID> ledger row ('written')
        "execution": {...},     # the dse_<ULID> candidate
        "import_contract_id": "cic_<ULID>",
        "no_op": False,
        "replay": bool,
        "row_count": int,
        "rejected_count": int,
        "blocked": False,
        "reason": None,
        "outcome": "written_pending_publication",  # HONEST: not yet published
        "published": False,                        # physical write + swap are Phase B
      }

    On an empty-blocked / rejection-blocked / no-op path the same keys are present with
    ``blocked``/``no_op`` set and ``published=False``.
    """
    # AC5: Append unavailable (before ANY parsing / ledger write).
    if write_mode == WRITE_MODE_APPEND:
        raise AppendUnavailable()

    if producer is not None:
        # File-source producer path (AD-1): the producer owns parsing + the
        # source->canonical mapping. feed_format stays a valid ledger enum
        # (csv/excel); the file-source template is the versioned contract.
        result = producer(data)
        fmt = source_metadata.get("format") or FORMAT_EXCEL
        if fmt not in SUPPORTED_FORMATS:
            fmt = FORMAT_EXCEL
    else:
        # Normalise and validate the contract.
        norm = validate_import_contract(contract)
        fmt = norm["format"]

        # H2: re-detect the actual format from the bytes and assert it matches the
        # contract (mirrors build_preview). A contract labelled csv over xlsx bytes
        # must NOT run parse_csv on binary -- fail with the invalid-contract error.
        detected = detect_format(source_metadata.get("filename"), data)
        if detected != fmt:
            raise InvalidImportContract(
                f"Contract format '{fmt}' does not match the actual file format "
                f"'{detected}' (detected from content/extension)."
            )

        # Parse (re-parse after the preview confirmation -- bytes are cheap).
        if fmt == FORMAT_CSV:
            result = parse_csv(
                data,
                delimiter=norm.get("delimiter"),
                date_format=norm.get("date_format"),
                locale=norm.get("locale"),
                header_row=norm.get("header_row", 1),
                column_types=norm.get("column_types"),
            )
        else:
            result = parse_excel(
                data,
                sheet_name=norm.get("sheet_name"),
                header_row=norm.get("header_row", 1),
                date_format=norm.get("date_format"),
                locale=norm.get("locale"),
                formulas_as_values=norm.get("formulas_as_values", True),
                column_types=norm.get("column_types"),
            )

    accepted_rows = result.rows
    rejected_rows_parsed = result.rejected
    content_hash = result.content_hash

    # AC4: Empty-file / zero-row guard (checked BEFORE opening the ledger so we never
    # mint a useless candidate for a trivially blocked import).
    allow_empty = (preferences or {}).get("allow_empty_publication", False)
    if len(accepted_rows) == 0:
        if not (force_empty_publish and allow_empty):
            return {
                "ledger": None,
                "execution": None,
                "import_contract_id": None,
                "no_op": False,
                "replay": False,
                "row_count": 0,
                "rejected_count": len(rejected_rows_parsed),
                "blocked": True,
                "reason": "empty_import",
                "outcome": "empty_import_blocked",
                "published": False,
            }

    # Import the module (not individual names) so tests can patch
    # ``core.managed_feed_ledger.*`` reliably.
    import core.managed_feed_ledger as _mfl  # noqa: PLC0415

    # AC2: version (or dedup) the parsing contract and thread its id onto the ledger.
    # File-source path (22.12): the versioned contract is the file-source template
    # (Story 22.11), not a CSV/Excel contract -- thread its id if the caller supplied
    # one; never version a CSV/Excel contract here.
    if producer is not None:
        import_contract_id = source_metadata.get("import_contract_id")
    else:
        import_contract_id = version_contract(
            contract,
            datastream_id=datastream_id,
            project_id=project_id,
            actor=actor,
            conn=conn,
        )

    # Delegate to the 12.8 ledger (open_import mints the ledger row + candidate BEFORE
    # any row is written -- NFR14).
    ledger_result = _mfl.open_import(
        datastream_id=datastream_id,
        project_id=project_id,
        plan_version_id=plan_version_id,
        mapping_version_id=mapping_version_id,
        feed_format=fmt,
        projection_plan=projection_plan,
        actor=actor,
        idempotency_key=idempotency_key,
        source_metadata={
            **source_metadata,
            "parsed_row_count": len(accepted_rows),
            "rejected_row_count": len(rejected_rows_parsed),
            "format": fmt,
        },
        content_hash=content_hash,
        conn=conn,
        write_mode=WRITE_MODE_REPLACE,
        import_contract_id=import_contract_id,
    )

    if ledger_result.get("no_op"):
        return {
            **ledger_result,
            "import_contract_id": import_contract_id,
            "row_count": 0,
            "rejected_count": 0,
            "blocked": False,
            "reason": None,
            "outcome": "noop",
            "published": False,
        }

    ledger = ledger_result["ledger"]
    ledger_id = ledger["id"]

    # Build the per-row rejection payloads for record_rows.
    rejection_detail = [
        {
            "row_number": rr.row_number,
            "field_name": rr.field_name,
            "rule": rr.rule,
            "reason": rr.reason,
            "rejected_value": _truncate_value(rr.rejected_value),
        }
        for rr in rejected_rows_parsed
    ]

    # Allocate the landing relation (one per datastream, project-scoped raw).
    landing_relation = _mfl.allocate_landing_relation(
        datastream_id=datastream_id,
        project_id=project_id,
        raw_schema=raw_schema,
    )

    # Phase B: the actual DuckDB/BigQuery row write via ``open_managed_landing`` goes
    # here (no live creds in dev). record_rows records the accepted count + the honest
    # per-row rejection evidence in the SAME transaction; the physical byte-for-byte
    # landing write is the Phase-B wiring.
    ledger_after = _mfl.record_rows(
        ledger_id=ledger_id,
        project_id=project_id,
        landing_relation=landing_relation,
        accepted_row_count=len(accepted_rows),
        content_hash=content_hash,
        rejected_rows=rejection_detail,
        actor=actor,
        conn=conn,
    )

    # Evaluate the UNMOCKED rejection gate against the ledger's recorded counts.
    gate_issue = _mfl.evaluate_rejection_gate_for_ledger(ledger_id, project_id, conn)
    if gate_issue:
        _mfl.mark_outcome(
            ledger_id=ledger_id,
            project_id=project_id,
            outcome=_mfl.OUTCOME_FAILED,
            actor=actor,
            error_code=gate_issue.get("code"),
            error_detail=gate_issue.get("detail"),
            conn=conn,
        )
        return {
            "ledger": ledger_after,
            "execution": ledger_result.get("execution"),
            "import_contract_id": import_contract_id,
            "no_op": False,
            "replay": ledger_result.get("replay", False),
            "row_count": len(accepted_rows),
            "rejected_count": len(rejected_rows_parsed),
            "blocked": True,
            "reason": "rejection_threshold_exceeded",
            "gate_issue": gate_issue,
            "outcome": _mfl.OUTCOME_FAILED,
            "published": False,
        }

    # Passing gate: the candidate is 'written' but NOT published. The physical write,
    # the 12.5 DQ gates, and the pointer-swap commit_publication are HONEST Phase B --
    # DO NOT claim a publish here.
    return {
        "ledger": ledger_after,
        "execution": ledger_result.get("execution"),
        "import_contract_id": import_contract_id,
        "no_op": False,
        "replay": ledger_result.get("replay", False),
        "row_count": len(accepted_rows),
        "rejected_count": len(rejected_rows_parsed),
        "blocked": False,
        "reason": None,
        "outcome": OUTCOME_WRITTEN_PENDING_PUBLICATION,
        "published": False,
    }
