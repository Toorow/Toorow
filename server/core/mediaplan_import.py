"""toorow -- Media plan Excel multi-sheet import parser + orchestration (Story 22.2).

Turns a real, messy media-plan workbook (multi-sheet, sub-totals, MERGED cells --
including a merged budget cell spanning several physical sub-rows) into a CANDIDATE
version through Story 22.1's already-shipped machinery
(``core.mediaplan_store.create_version_with_lines``). Publication stays an explicit
second step (the 22.1 publish API): an import is NEVER destructive -- the currently
published version stays intact whatever happens, and no candidate is created when
zero valid line survives parsing.

Merged-cell policy (exigence Jean 2026-07-21 -- OBLIGATOIRE):
  (a) merged ranges are RESOLVED before parsing: every covered cell knows the
      top-left value and the range id of its merge;
  (b) merged DESCRIPTIVE fields (label/channel/dates) forward-fill onto every
      physical row the merge covers -- bounded to the range, never beyond;
  (c) a merged BUDGET cell over N physical rows collapses the group into ONE plan
      line (one budget, one span). NEVER a forward-fill of the budget (that would
      multiply it by N = double-count) and NEVER a silent split. Exploding the
      group into N lines is an EXPLICIT contract choice
      (``explode_merged_budget`` with a caller-supplied split that must sum EXACTLY
      to the merged budget, else 422). ``explode_merged_budget`` is keyed by the
      PHYSICAL range id (e.g. "E2:E4"): the explode binds to the merge's physical
      A1 coordinates. If next month's file shifts the merge (different rows, hence
      a different range_id), the explode NO LONGER applies and the group collapses
      to ONE line by default -- it is NEVER silently exploded on a stale key;
  (d) the invariant holds and is PROVEN in the report. ``sum_file`` is computed by
      an INDEPENDENT pass over the budget column (OUT of the normalisation path):
      every distinct merged-budget range counts ONCE (top-left value), every
      unmerged numeric budget cell counts once -- including cells on rows that will
      later be rejected for OTHER reasons (bad dates, etc.). The parser then asserts
      ``sum_file == sum_imported + sum_rejected`` (sum_rejected = the readable
      budgets carried by rejected rows, surfaced to the report). A double-count of a
      merge makes this diverge and raises loudly; a rejected-but-readable budget is
      visible in the report, NEVER melted silently.

Business errors reuse the typed MediaPlan*Error family from Story 22.1; the API
layer maps them to 4xx and never leaks str(exc) into a 5xx body (lesson 12.3).

Convergence with the generic Epic 12 import contracts (FR30/FR31) is intentionally
OUT OF SCOPE (orchestration decision Jean 2026-07-21): this parser stands on 22.1's
store, not on the Epic 12 managed-feed writer.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from core.audit import ACTION_MEDIA_PLAN_IMPORTED, insert_audit_row
from core.mediaplan_store import (
    _CENT,
    MediaPlanNotFoundError,
    MediaPlanValidationError,
    _parse_budget,
    create_version_with_lines,
    list_lines,
)

# Target fields a contract column may map to. line_key is handled separately
# (contract["line_key"]), the rest are line attributes.
_TARGET_FIELDS = frozenset(
    {
        "line_key",
        "label",
        "channel",
        "start_date",
        "end_date",
        "budget",
        "buy_mode",
        "is_plan_only",
    }
)
# Descriptive fields that forward-fill across a merge range (rule b).
_DESCRIPTIVE_FIELDS = frozenset({"label", "channel", "start_date", "end_date", "buy_mode"})

# DoS guards (F-5): reject absurdly large sheets BEFORE any per-cell loop. A real
# media-plan sheet is far below these bounds; anything above is either a crafted
# bomb or a corrupt file, never a legitimate import.
MAX_SHEET_ROWS = 100_000
MAX_SHEET_COLUMNS = 200


class MediaPlanImportError(MediaPlanValidationError):
    """A workbook could not be parsed (unreadable/empty/contract mismatch) -> 422."""

    code = "import_error"


@dataclass
class ImportPreview:
    """Result of parsing a workbook against its per-sheet contracts.

    ``lines`` are normalised payloads ready for create_version_with_lines. Each
    line may carry a ``detail`` list (the sub-line labels of a merged-budget
    group), transported in the report but NOT persisted in v1 (documented).
    """

    lines: list[dict[str, Any]] = field(default_factory=list)
    rejected: list[dict[str, Any]] = field(default_factory=list)
    sum_file: Decimal = Decimal("0.00")
    sum_imported: Decimal = Decimal("0.00")
    sum_rejected: Decimal = Decimal("0.00")


# ---------------------------------------------------------------------------
# slugify (line_key auto composition)
# ---------------------------------------------------------------------------


def slugify(value: str) -> str:
    """ASCII-ish slug for line_key composition (sheet/label).

    Lowercase, accents folded to ASCII, non-alphanumerics collapsed to '-'.
    Kept dependency-free (no external transliteration lib).
    """
    import unicodedata

    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    out: list[str] = []
    prev_dash = False
    for ch in text:
        if ch.isalnum():
            out.append(ch)
            prev_dash = False
        elif not prev_dash:
            out.append("-")
            prev_dash = True
    return "".join(out).strip("-") or "ligne"


# ---------------------------------------------------------------------------
# Value coercion (dates / budgets) -- reuse 22.1 guards where possible
# ---------------------------------------------------------------------------


def _parse_cell_date(value: Any, *, date_format: str | None) -> date | None:
    """Coerce an Excel cell to a date, or None if empty/unparseable.

    Accepts: native datetime/date (openpyxl returns datetime for date cells),
    ISO strings, and an optional contract date_format (e.g. %d/%m/%Y). Returns
    None on empty so the caller can decide whether the row is noise.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    if date_format:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            pass
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_cell_budget(value: Any) -> Decimal | None:
    """Coerce a budget cell to a cent-normalised Decimal, or None if empty.

    Handles native numbers and FR-formatted strings ("1 234,56", with regular,
    thin (\\u202f/\\u00a0) spaces and a '.' or ',' decimal separator). Reuses the
    22.1 _parse_budget guard for the final validation (finite, non-negative,
    quantised to the cent). A non-empty but non-numeric cell returns None (the
    caller treats it as "no numeric budget" -> noise row).
    """
    if value is None:
        return None
    if isinstance(value, bool):  # bool is an int subclass; never a budget
        return None
    if isinstance(value, (int, float, Decimal)):
        return _parse_budget(value)
    text = str(value).strip()
    if not text:
        return None
    # Strip ALL Unicode whitespace (incl. NBSP U+00A0 and narrow NBSP U+202F used
    # as FR thousands separators) and currency symbols; normalise the FR decimal
    # comma. Done char-by-char so no exotic space literal lives in the source.
    cleaned = "".join(ch for ch in text if not ch.isspace())
    for _token in ("\N{EURO SIGN}", "EUR", "eur"):
        cleaned = cleaned.replace(_token, "")
    if "," in cleaned and "." in cleaned:
        # Ambiguous: assume '.' thousands + ',' decimal (FR) -> drop dots, comma->dot.
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return _parse_budget(cleaned)
    except MediaPlanValidationError:
        return None


# ---------------------------------------------------------------------------
# Merge resolution
# ---------------------------------------------------------------------------


def _resolve_merges(ws: Any) -> dict[tuple[int, int], tuple[Any, str]]:
    """Build {(row, col) -> (top_left_value, range_id)} for every merged cell.

    row/col are 1-based (openpyxl convention). range_id is the A1 range string
    (e.g. "C4:C6"), stable and human-readable -- used as the explode key and in
    rejection reasons.
    """
    covered: dict[tuple[int, int], tuple[Any, str]] = {}
    for rng in ws.merged_cells.ranges:
        range_id = str(rng)
        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                covered[(r, c)] = (top_left, range_id)
    return covered


def _sheet_file_budget_sum(
    ws: Any,
    merges: dict[tuple[int, int], tuple[Any, str]],
    *,
    budget_col: int,
    first_data_row: int,
    last_row: int,
) -> Decimal:
    """Independent pass (F-1): sum every readable numeric budget of the data zone.

    Computed OUTSIDE the normalisation path so it can never tautologically equal
    ``sum_imported``. Each distinct merged-budget range counts ONCE (its top-left
    value, by range_id); each unmerged numeric budget cell counts once. Cells on
    rows that will later be rejected for OTHER reasons (bad dates, etc.) ARE counted
    -- honesty over the whole file. The invariant then asserts this equals
    sum_imported + sum_rejected (a merge double-count diverges; a readable-but-
    rejected budget stays visible, never melted).
    """
    total = Decimal("0.00")
    seen_ranges: set[str] = set()
    for row in range(first_data_row, last_row + 1):
        raw, budget_range = _cell_resolved(ws, merges, row, budget_col)
        if budget_range is not None:
            if budget_range in seen_ranges:
                continue  # a merged budget counts ONCE (top-left), not per row
            seen_ranges.add(budget_range)
        budget = _parse_cell_budget(raw)
        if budget is not None:
            total += budget
    return total


def _cell_resolved(
    ws: Any, merges: dict[tuple[int, int], tuple[Any, str]], row: int, col: int
) -> tuple[Any, str | None]:
    """Return (value, range_id|None) for a cell, resolving merges (rule a).

    A covered cell reads the top-left value and its range id; an unmerged cell
    reads its own value and None.
    """
    if (row, col) in merges:
        return merges[(row, col)]
    return ws.cell(row=row, column=col).value, None


# ---------------------------------------------------------------------------
# Column resolution (header text or letter -> field)
# ---------------------------------------------------------------------------


def _column_index(ws: Any, header_row: int, key: str) -> int | None:
    """Resolve a contract column key to a 1-based column index.

    ``key`` is either a header cell text (matched case-insensitively, trimmed,
    against the header row) or an Excel column letter (e.g. "C"). Header text is
    tried FIRST so a header that happens to read like a letter (e.g. "TV") is not
    mis-parsed as a column reference; a pure-letter key is used as a fallback only
    when no header matches. Returns None when unresolved.
    """
    from openpyxl.utils import column_index_from_string  # noqa: PLC0415

    stripped = key.strip()
    target = stripped.casefold()
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None and str(val).strip().casefold() == target:
            return col
    # Fallback: interpret the key as an Excel column letter (A, AB, ...).
    if stripped.isalpha() and stripped.isupper():
        try:
            return column_index_from_string(stripped)
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Contract validation
# ---------------------------------------------------------------------------


def validate_contract(contract: dict[str, Any]) -> dict[str, Any]:
    """Validate a single-sheet contract; return a normalised copy.

    Raises MediaPlanValidationError (422) on: non-int/<1 header_row, empty columns,
    unknown target field, missing budget mapping, bad explode split shape.
    """
    if not isinstance(contract, dict):
        raise MediaPlanValidationError("Le contrat doit être un objet.")

    header_row = contract.get("header_row")
    if not isinstance(header_row, int) or isinstance(header_row, bool) or header_row < 1:
        raise MediaPlanValidationError(
            "Le contrat doit préciser « header_row » (entier >= 1, base 1)."
        )

    columns = contract.get("columns")
    if not isinstance(columns, dict) or not columns:
        raise MediaPlanValidationError(
            "Le contrat doit préciser un mapping « columns » non vide."
        )
    fields = set(columns.values())
    unknown = fields - _TARGET_FIELDS
    if unknown:
        raise MediaPlanValidationError(
            "Champs cibles inconnus dans le contrat : "
            + ", ".join(sorted(str(u) for u in unknown))
            + "."
        )
    if "budget" not in fields:
        raise MediaPlanValidationError(
            "Le contrat doit mapper une colonne vers le champ « budget »."
        )

    line_key = contract.get("line_key", "auto")
    if not isinstance(line_key, str) or not line_key.strip():
        raise MediaPlanValidationError(
            "Le contrat doit préciser « line_key » (« auto » ou un nom de colonne)."
        )

    explode = contract.get("explode_merged_budget", {})
    if explode is None:
        explode = {}
    if not isinstance(explode, dict):
        raise MediaPlanValidationError(
            "« explode_merged_budget » doit être un objet {range_id: [poids...]}."
        )
    for range_id, weights in explode.items():
        if not isinstance(weights, list) or not weights:
            raise MediaPlanValidationError(
                f"« explode_merged_budget[{range_id}] » doit être une liste de poids non vide."
            )
        # F-6: validate each weight is numeric and strictly positive at the PUT,
        # not only at import time. Bool is an int subclass -> explicitly rejected.
        for w in weights:
            if isinstance(w, bool) or not isinstance(w, (int, float, str)):
                raise MediaPlanValidationError(
                    f"« explode_merged_budget[{range_id}] » : poids « {w} » invalide "
                    "(nombre attendu)."
                )
            try:
                weight = Decimal(str(w))
            except (ArithmeticError, ValueError):
                raise MediaPlanValidationError(
                    f"« explode_merged_budget[{range_id}] » : poids « {w} » invalide "
                    "(nombre attendu)."
                ) from None
            if not weight.is_finite() or weight <= 0:
                raise MediaPlanValidationError(
                    f"« explode_merged_budget[{range_id}] » : répartition invalide "
                    f"(poids négatif ou nul « {w} »)."
                )

    date_format = contract.get("date_format")
    if date_format is not None and not isinstance(date_format, str):
        raise MediaPlanValidationError("« date_format » doit être une chaîne (ex. %d/%m/%Y).")

    return {
        "header_row": header_row,
        "columns": {str(k): str(v) for k, v in columns.items()},
        "line_key": line_key.strip(),
        "date_format": date_format,
        "is_plan_only": bool(contract.get("is_plan_only", False)),
        "explode_merged_budget": explode,
    }


# ---------------------------------------------------------------------------
# Per-sheet parsing
# ---------------------------------------------------------------------------


def _parse_sheet(
    ws: Any,
    contract: dict[str, Any],
    *,
    sheet_name: str,
    seen_keys: dict[str, str],
    preview: ImportPreview,
) -> None:
    """Parse one contractualised sheet into ``preview`` (mutates it).

    Groups physical rows by a merged BUDGET range (rule c): a merged budget cell
    collapses its covered rows into ONE line unless the contract explicitly
    explodes that range. Descriptive fields forward-fill per merge range (rule b).
    Noise rows (no valid dates OR no numeric budget, not covered by a budget
    merge) are rejected with a precise reason.
    """
    # DoS guard (F-5): bound the sheet dimensions BEFORE any per-cell loop.
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row > MAX_SHEET_ROWS:
        raise MediaPlanImportError(
            f"Onglet « {sheet_name} » : feuille trop grande "
            f"({max_row} lignes, maximum {MAX_SHEET_ROWS})."
        )
    if max_col > MAX_SHEET_COLUMNS:
        raise MediaPlanImportError(
            f"Onglet « {sheet_name} » : feuille trop grande "
            f"({max_col} colonnes, maximum {MAX_SHEET_COLUMNS})."
        )

    merges = _resolve_merges(ws)
    header_row = contract["header_row"]
    columns = contract["columns"]
    date_format = contract["date_format"]
    explode = contract["explode_merged_budget"]
    contract_plan_only = contract["is_plan_only"]

    # Resolve each mapped field to a column index (once).
    field_col: dict[str, int] = {}
    for key, tgt in columns.items():
        idx = _column_index(ws, header_row, key)
        if idx is None:
            raise MediaPlanImportError(
                f"Onglet « {sheet_name} » : colonne « {key} » introuvable "
                f"(ligne d'en-tête {header_row})."
            )
        field_col[tgt] = idx
    budget_col = field_col["budget"]

    # A contract may also name the line_key as a column (not "auto").
    lk_mode = contract["line_key"]
    lk_col: int | None = None
    if lk_mode != "auto":
        lk_col = _column_index(ws, header_row, lk_mode)
        if lk_col is None:
            raise MediaPlanImportError(
                f"Onglet « {sheet_name} » : colonne line_key « {lk_mode} » introuvable."
            )

    first_data_row = header_row + 1
    last_row = ws.max_row or header_row

    # F-1: file-side budget total by an INDEPENDENT pass over the budget column,
    # accumulated across sheets. Never derived from the normalisation path below,
    # so the final invariant can actually catch a merge double-count.
    preview.sum_file += _sheet_file_budget_sum(
        ws, merges, budget_col=budget_col,
        first_data_row=first_data_row, last_row=last_row,
    )

    def _field_value(row: int, tgt: str) -> tuple[Any, str | None]:
        col = field_col[tgt]
        return _cell_resolved(ws, merges, row, col)

    # Walk data rows, grouping consecutive rows that share a merged budget range.
    row = first_data_row
    while row <= last_row:
        raw_budget, budget_range = _cell_resolved(ws, merges, row, budget_col)

        if budget_range is None:
            _parse_single_row(
                ws, merges, contract, sheet_name=sheet_name, row=row,
                field_col=field_col, lk_col=lk_col, date_format=date_format,
                contract_plan_only=contract_plan_only, seen_keys=seen_keys,
                preview=preview, field_value=_field_value,
            )
            row += 1
            continue

        # Merged budget: gather the full run of rows sharing this range.
        group_rows: list[int] = []
        r = row
        while r <= last_row:
            _, rng = _cell_resolved(ws, merges, r, budget_col)
            if rng != budget_range:
                break
            group_rows.append(r)
            r += 1

        _parse_budget_group(
            ws, merges, contract, sheet_name=sheet_name, group_rows=group_rows,
            budget_range=budget_range, raw_budget=raw_budget, explode=explode,
            field_col=field_col, lk_col=lk_col, date_format=date_format,
            contract_plan_only=contract_plan_only, seen_keys=seen_keys,
            preview=preview, field_value=_field_value,
        )
        row = r


def _compose_line_key(
    *, lk_col: int | None, ws: Any, merges: dict, row: int, sheet_name: str, label: str
) -> str:
    if lk_col is not None:
        val, _ = _cell_resolved(ws, merges, row, lk_col)
        key = str(val or "").strip()
        if key:
            return key
    return f"{slugify(sheet_name)}/{slugify(label)}"


def _register_key(seen_keys: dict[str, str], key: str, sheet_name: str) -> None:
    if key in seen_keys:
        raise MediaPlanValidationError(
            f"Clé de ligne dupliquée « {key} » "
            f"(onglets « {seen_keys[key]} » et « {sheet_name} »)."
        )
    seen_keys[key] = sheet_name


def _parse_single_row(
    ws, merges, contract, *, sheet_name, row, field_col, lk_col, date_format,
    contract_plan_only, seen_keys, preview, field_value,
) -> None:
    """Parse one unmerged-budget physical row (rule: noise exclusion)."""
    raw_budget, _ = field_value(row, "budget")
    budget = _parse_cell_budget(raw_budget)

    start = _read_date(field_value, row, "start_date", date_format)
    end = _read_date(field_value, row, "end_date", date_format)
    label = _read_text(field_value, row, "label")

    if budget is None:
        preview.rejected.append(
            {"row": row, "sheet": sheet_name, "label": label,
             "raison": "ligne ignorée : pas de budget numérique (sous-total ?)."}
        )
        return
    if start is None or end is None:
        # Rejected but the budget was READABLE: count it in sum_rejected so the
        # invariant (sum_file == sum_imported + sum_rejected) still holds and the
        # amount is surfaced, never melted silently (F-1).
        preview.rejected.append(
            {"row": row, "sheet": sheet_name, "label": label,
             "budget": format(budget, "f"),
             "raison": "ligne ignorée : pas de dates valides."}
        )
        preview.sum_rejected += budget
        return

    line = _build_line(
        ws=ws, merges=merges, row=row, sheet_name=sheet_name, label=label,
        budget=budget, start=start, end=end, field_value=field_value,
        contract_plan_only=contract_plan_only, lk_col=lk_col,
    )
    _register_key(seen_keys, line["line_key"], sheet_name)
    preview.lines.append(line)
    preview.sum_imported += budget


def _parse_budget_group(
    ws, merges, contract, *, sheet_name, group_rows, budget_range, raw_budget,
    explode, field_col, lk_col, date_format, contract_plan_only, seen_keys,
    preview, field_value,
) -> None:
    """Parse a merged-budget group (rule c): ONE line, or explicit explode."""
    group_budget = _parse_cell_budget(raw_budget)
    if group_budget is None:
        # A merged, non-numeric budget block is noise (e.g. a merged header band).
        for r in group_rows:
            preview.rejected.append(
                {"row": r, "sheet": sheet_name,
                 "raison": f"groupe fusionné {budget_range} ignoré : budget non numérique."}
            )
        return

    # Gather per-row descriptive data (dates + sub-labels).
    sub: list[dict[str, Any]] = []
    for r in group_rows:
        sub.append(
            {
                "row": r,
                "label": _read_text(field_value, r, "label"),
                "start": _read_date(field_value, r, "start_date", date_format),
                "end": _read_date(field_value, r, "end_date", date_format),
            }
        )

    if budget_range in explode:
        _explode_group(
            ws=ws, merges=merges, sheet_name=sheet_name, budget_range=budget_range,
            group_budget=group_budget, weights=explode[budget_range], sub=sub,
            field_value=field_value, contract_plan_only=contract_plan_only,
            lk_col=lk_col, seen_keys=seen_keys, preview=preview,
        )
        return

    # Default: collapse into ONE plan line (never multiply, never silent split).
    starts = [s["start"] for s in sub if s["start"] is not None]
    ends = [s["end"] for s in sub if s["end"] is not None]
    if not starts or not ends:
        # Rejected group, but its (single) merged budget was READABLE: count it in
        # sum_rejected ONCE so the invariant holds and the amount stays visible (F-1).
        for r in group_rows:
            preview.rejected.append(
                {"row": r, "sheet": sheet_name,
                 "raison": f"groupe fusionné {budget_range} ignoré : aucune date valide."}
            )
        preview.sum_rejected += group_budget
        return

    start = min(starts)
    end = max(ends)
    # Label: the merged label value if the label itself is merged over the group,
    # else the first sub-label + a "(groupe de N)" suffix.
    label_val, label_range = field_value(group_rows[0], "label")
    covers_group = label_range is not None and all(
        _cell_resolved(ws, merges, r, field_col["label"])[1] == label_range
        for r in group_rows
    )
    if covers_group and label_val:
        label = str(label_val).strip()
    else:
        first = next((s["label"] for s in sub if s["label"]), "")
        n = len(group_rows)
        label = f"{first} (groupe de {n})" if first else f"Groupe de {n}"

    detail = [s["label"] for s in sub if s["label"]]
    line = _build_line(
        ws=ws, merges=merges, row=group_rows[0], sheet_name=sheet_name, label=label,
        budget=group_budget, start=start, end=end, field_value=field_value,
        contract_plan_only=contract_plan_only, lk_col=lk_col,
    )
    # detail: sub-line labels transported in the report, NOT persisted in v1.
    line["detail"] = detail
    _register_key(seen_keys, line["line_key"], sheet_name)
    preview.lines.append(line)
    preview.sum_imported += group_budget


def _explode_group(
    ws, merges, *, sheet_name, budget_range, group_budget, weights, sub,
    field_value, contract_plan_only, lk_col, seen_keys, preview,
) -> None:
    """Explicit explode: N lines with a caller split that must sum EXACTLY (422)."""
    if len(weights) != len(sub):
        raise MediaPlanValidationError(
            f"Éclatement de {budget_range} : {len(weights)} poids fournis pour "
            f"{len(sub)} sous-lignes."
        )
    parts: list[Decimal] = []
    for w in weights:
        try:
            weight = Decimal(str(w))
        except Exception as exc:  # noqa: BLE001
            raise MediaPlanValidationError(
                f"Éclatement de {budget_range} : poids « {w} » invalide."
            ) from exc
        # F-2: a negative or null weight is rejected immediately with a precise
        # message (guarded here too, not only at contract validation time).
        if not weight.is_finite() or weight <= 0:
            raise MediaPlanValidationError(
                f"Éclatement de {budget_range} : répartition invalide "
                f"(poids négatif ou nul « {w} »)."
            )
        parts.append(weight.quantize(_CENT))
    total = sum(parts, Decimal("0.00"))
    if total != group_budget:
        raise MediaPlanValidationError(
            f"Éclatement de {budget_range} : la répartition ({total}) ne somme pas "
            f"exactement au budget fusionné ({group_budget})."
        )

    for part, s in zip(parts, sub):
        if s["start"] is None or s["end"] is None:
            raise MediaPlanValidationError(
                f"Éclatement de {budget_range} : la sous-ligne « {s['label']} » "
                f"n'a pas de dates valides."
            )
        line = _build_line(
            ws=ws, merges=merges, row=s["row"], sheet_name=sheet_name,
            label=s["label"] or f"{budget_range} sous-ligne", budget=part,
            start=s["start"], end=s["end"], field_value=field_value,
            contract_plan_only=contract_plan_only, lk_col=lk_col,
        )
        _register_key(seen_keys, line["line_key"], sheet_name)
        preview.lines.append(line)
        preview.sum_imported += part


def _build_line(
    *, ws, merges, row, sheet_name, label, budget, start, end, field_value,
    contract_plan_only, lk_col,
) -> dict[str, Any]:
    channel = _read_text(field_value, row, "channel") or None
    buy_mode = _read_text(field_value, row, "buy_mode") or None
    # is_plan_only default comes from the contract; a mapped column overrides it.
    is_plan_only = contract_plan_only
    if _has_field(field_value, row, "is_plan_only"):
        raw_po, _ = field_value(row, "is_plan_only")
        if raw_po is not None:
            is_plan_only = _coerce_bool(raw_po)

    if start > end:
        raise MediaPlanValidationError(
            f"Onglet « {sheet_name} », ligne {row} : date de début postérieure à la fin."
        )
    line_key = _compose_line_key(
        lk_col=lk_col, ws=ws, merges=merges, row=row, sheet_name=sheet_name, label=label
    )
    return {
        "line_key": line_key,
        "label": label or "",
        "channel": channel,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "budget": format(budget, "f"),
        "buy_mode": buy_mode,
        "is_plan_only": bool(is_plan_only),
    }


# -- small read helpers over the field_value closure -------------------------


def _read_text(field_value, row: int, tgt: str) -> str:
    if not _has_field(field_value, row, tgt):
        return ""
    val, _ = field_value(row, tgt)
    return str(val).strip() if val is not None else ""


def _read_date(field_value, row: int, tgt: str, date_format: str | None) -> date | None:
    if not _has_field(field_value, row, tgt):
        return None
    val, _ = field_value(row, tgt)
    return _parse_cell_date(val, date_format=date_format)


def _has_field(field_value, row: int, tgt: str) -> bool:
    try:
        field_value(row, tgt)
        return True
    except KeyError:
        return False


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value).strip().casefold()
    return text in {"1", "true", "vrai", "oui", "yes", "x", "plan-only", "plan only"}


# ---------------------------------------------------------------------------
# Workbook entrypoint
# ---------------------------------------------------------------------------


def parse_workbook(
    file_bytes: bytes, contracts: dict[str, dict[str, Any]]
) -> ImportPreview:
    """Parse an xlsx (bytes) against per-sheet contracts -> ImportPreview.

    Only sheets present in ``contracts`` are parsed. Raises MediaPlanImportError
    on unreadable/empty files or a contract naming a missing sheet; raises
    MediaPlanValidationError on business issues (bad contract, duplicate line_key,
    bad explode split). The final invariant sum_file == sum_imported + sum_rejected
    is asserted -- sum_file being computed by an INDEPENDENT pass (F-1), a mismatch
    means a merge was double-counted: a parser bug, surfaced loudly.
    """
    if not file_bytes:
        raise MediaPlanImportError("Fichier vide : aucun onglet à importer.")
    if not contracts:
        raise MediaPlanImportError("Aucun contrat d'import : rien à parser.")

    try:
        from openpyxl import load_workbook  # noqa: PLC0415

        # read_only=False so merged_cells ranges are available; data_only=True so
        # formula cells yield their last computed value, not the formula string.
        wb = load_workbook(
            io.BytesIO(file_bytes), read_only=False, data_only=True
        )
    except MediaPlanImportError:
        raise
    except Exception as exc:  # noqa: BLE001
        # Corrupt / non-xlsx bytes: a clean typed error, never str(exc) to client.
        raise MediaPlanImportError(
            "Fichier illisible : ce n'est pas un classeur Excel (.xlsx) valide."
        ) from exc

    normalised = {name: validate_contract(c) for name, c in contracts.items()}

    preview = ImportPreview()
    seen_keys: dict[str, str] = {}
    for sheet_name, contract in normalised.items():
        if sheet_name not in wb.sheetnames:
            raise MediaPlanImportError(
                f"Onglet « {sheet_name} » absent du classeur."
            )
        ws = wb[sheet_name]
        _parse_sheet(
            ws, contract, sheet_name=sheet_name, seen_keys=seen_keys, preview=preview
        )

    # Invariant (rule d, F-1): the INDEPENDENT file total equals what was imported
    # plus what was rejected-but-readable. sum_file is NOT derived from the
    # normalisation path, so a merge double-counted below makes this diverge.
    if preview.sum_file != preview.sum_imported + preview.sum_rejected:
        raise MediaPlanImportError(
            "Invariant d'import rompu : la somme du fichier "
            f"({preview.sum_file}) diffère de la somme importée + rejetée "
            f"({preview.sum_imported} + {preview.sum_rejected}). "
            "Bug du parseur (double-compte de fusion ?), jamais silencieux."
        )
    return preview


# ---------------------------------------------------------------------------
# Orchestration: contract load + parse + candidate creation + report
# ---------------------------------------------------------------------------


def load_contracts(conn: Any, *, plan_id: str) -> dict[str, dict[str, Any]]:
    """Load every import contract of a plan as {sheet_name: contract jsonb}."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT sheet_name, contract
            FROM app.media_plan_import_contracts
            WHERE plan_id = %s
            ORDER BY sheet_name
            """,
            (plan_id,),
        )
        return {row[0]: row[1] for row in cur.fetchall()}


def _active_lines_by_key(conn: Any, *, plan_id: str) -> dict[str, dict[str, Any]]:
    """Map line_key -> the active version's line (empty when no active version)."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id FROM app.media_plan_versions WHERE plan_id = %s AND is_active",
            (plan_id,),
        )
        row = cur.fetchone()
    if not row:
        return {}
    return {ln["line_key"]: ln for ln in list_lines(conn, version_id=row[0])}


def _classify(
    new_line: dict[str, Any], active: dict[str, dict[str, Any]]
) -> str:
    """created / updated / unchanged vs the active version, by line_key value-compare."""
    prev = active.get(new_line["line_key"])
    if prev is None:
        return "created"
    for f in ("label", "channel", "start_date", "end_date", "budget", "buy_mode"):
        pv = prev.get(f)
        nv = new_line.get(f)
        # buy_mode/channel may be None both sides.
        if (pv or None) != (nv or None):
            return "updated"
    if bool(prev.get("is_plan_only")) != bool(new_line.get("is_plan_only")):
        return "updated"
    return "unchanged"


def import_workbook(
    conn: Any, *, plan_id: str, file_bytes: bytes, actor: str
) -> dict[str, Any]:
    """Import an xlsx into a CANDIDATE version + return the import report.

    Two-step by design: this creates a candidate (never publishes). The published
    version stays intact whatever happens; when zero valid line survives parsing,
    NO candidate is created (import is never destructive). Publication is an
    explicit second step (POST /mediaplans/versions/{id}/publish, Story 22.1).

    Report shape:
      {version, per_line: {created, updated, unchanged: [line_key...]},
       rejected: [...], sums: {file, imported, rejected}}
    """
    with conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM app.media_plans WHERE id = %s AND archived_at IS NULL",
            (plan_id,),
        )
        if cur.fetchone() is None:
            raise MediaPlanNotFoundError("Plan introuvable.")

    contracts = load_contracts(conn, plan_id=plan_id)
    if not contracts:
        raise MediaPlanValidationError(
            "Aucun contrat d'import défini pour ce plan : "
            "configurez d'abord un contrat par onglet."
        )

    preview = parse_workbook(file_bytes, contracts)
    if not preview.lines:
        # Nothing valid: never touch the active version, never create an empty
        # candidate. Surface the rejections so the operator sees WHY.
        raise MediaPlanImportError(
            "Aucune ligne valide dans le fichier : la version publiée reste intacte "
            f"({len(preview.rejected)} ligne(s) rejetée(s))."
        )

    active = _active_lines_by_key(conn, plan_id=plan_id)
    per_line: dict[str, list[str]] = {"created": [], "updated": [], "unchanged": []}
    for line in preview.lines:
        per_line[_classify(line, active)].append(line["line_key"])

    version = create_version_with_lines(
        conn,
        plan_id=plan_id,
        lines=preview.lines,
        source_note="import Excel",
        created_by=actor,
    )

    insert_audit_row(
        conn,
        identity=actor,
        action=ACTION_MEDIA_PLAN_IMPORTED,
        provider_account="platform",
        connection_ref="",
        metadata={
            "plan_id": plan_id,
            "version_id": version["id"],
            "line_count": len(preview.lines),
            "rejected_count": len(preview.rejected),
            "sum_file": format(preview.sum_file, "f"),
            "sum_rejected": format(preview.sum_rejected, "f"),
        },
    )

    # Attach the per-line detail (merged-group sub-labels) to the report, keyed by
    # line_key (detail is transported here, NOT persisted in v1).
    details = {
        ln["line_key"]: ln["detail"] for ln in preview.lines if ln.get("detail")
    }
    return {
        "version": version,
        "per_line": per_line,
        "rejected": preview.rejected,
        "detail": details,
        "sums": {
            "file": format(preview.sum_file, "f"),
            "imported": format(preview.sum_imported, "f"),
            "rejected": format(preview.sum_rejected, "f"),
        },
    }
