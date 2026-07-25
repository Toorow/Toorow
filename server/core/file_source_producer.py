"""toorow -- Canonical-rows producer: reshape a workbook to DAILY rows (Story 22.10).

Phase B-0 of the file-source ingestion framework. Consumes the shared reshape
engine (``core.reshape``, Story 22.9 / AD-5) to turn a real, messy media-plan
workbook -- a top ``label: value`` metadata block, a header that is NOT row 1,
grouped lines with subtotal/total rows, merged cells, and per-line Start/Ende
date ranges -- into standardized DAILY canonical rows, output as the existing
``ParseResult`` (Story 12.9). It is the reshape core of the B-1 producer; the
catalog/template wiring at ``csv_excel_import.run_import`` is Story 22.12.

Architecture boundaries (AD-4 / AD-5):
  * ONE Excel engine: merged-cell resolution, header-row column resolution, and
    the cent-exact daily spread come from ``core.reshape`` -- this module never
    reimplements them, and no second Excel parser is introduced.
  * Ingestion-only: output is canonical RAW rows at the DAILY grain (+ rejected
    rows). No metric, FX, money composition, or analytics -- those stay in dbt.
  * No silent drop: group-label / subtotal / total / dateless rows become honest
    ``RejectedRow`` evidence with a stable rule code, never dropped in silence.

Sum-preserving explode (AD-5): each accepted line's amount is spread across its
[start, end] date range by ``reshape.compute_spread`` (largest-remainder at the
cent), so ``SUM(daily amounts) == line total`` EXACTLY, to the cent. The producer
asserts this per line defensively (a mismatch is a bug, surfaced loudly).

The producer is a PURE parse layer (no I/O, unit-testable on in-memory bytes),
mirroring ``csv_excel_import``'s parse layer. Landing / ledger / publication is
the caller's concern (Story 22.12 reuses ``managed_feed_ledger``).
"""

from __future__ import annotations

import hashlib
import io
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any

from core.csv_excel_import import (
    ColumnSpec,
    CsvExcelImportError,
    ParseResult,
    RejectedRow,
)
from core.file_source_template import PLACEMENT_CLASSES
from core.reshape import CENT, cell_resolved, compute_spread, resolve_column_index, resolve_merges

# Reserved per-row key stamping the matrix class (planned/actual/extrapolated).
# Underscore-prefixed so it never collides with a canonical mdm_ field id.
PLACEMENT_CLASS_KEY = "_placement_class"

# ---------------------------------------------------------------------------
# DoS guards (bound the sheet BEFORE any per-cell loop; same spirit as the
# csv_excel_import / mediaplan_import limits).
# ---------------------------------------------------------------------------

MAX_SHEET_ROWS = 500_000
MAX_SHEET_COLUMNS = 500

# How far down we scan for a header row when it is auto-detected.
HEADER_SCAN_LIMIT = 200

# ---------------------------------------------------------------------------
# Per-row rejection rule codes (stable; carried on RejectedRow).
# ---------------------------------------------------------------------------

RULE_SUBTOTAL_OR_TOTAL = "subtotal_or_total"  # a group subtotal / grand total row
RULE_GROUP_OR_NOISE = "group_or_noise"  # a group-label / blank row (no amount, no dates)
RULE_MISSING_DATES = "missing_dates"  # an amount is present but the date range is absent
RULE_BAD_AMOUNT = "bad_amount"  # the amount cell is non-numeric where a line is expected
RULE_INVERTED_DATES = "inverted_dates"  # start_date > end_date


class ReshapeProducerError(CsvExcelImportError):
    """A workbook could not be reshaped (unreadable / no header / no data) -> 422.

    Subclasses the shared import error so the API maps it like every other
    import-layer failure (stable ``code`` + optional ``repair``).
    """

    def __init__(self, code: str, detail: str = "", *, repair: dict | None = None) -> None:
        super().__init__(code, detail, repair=repair)


# ---------------------------------------------------------------------------
# Reshape spec (what the producer needs to reshape one sheet to daily rows).
#
# Story 22.10 keeps the spec explicit/declared -- tolerant recognition (B-2) and
# the versioned template artifact (B-1) resolve this spec for the caller later.
# ---------------------------------------------------------------------------


@dataclass
class ReshapeSpec:
    """Declares how to reshape one sheet's lines into daily canonical rows.

    ``fields`` maps canonical field id -> source column key (a header cell text,
    matched case-insensitively, or an Excel column letter as a fallback). It MUST
    include the amount, start, and end columns (named by ``amount_field`` /
    ``start_field`` / ``end_field``). On output the start/end columns are consumed
    (replaced by the single ``date_field`` grain); every other mapped field is
    carried onto each daily row, and ``amount_field`` carries the DAILY amount.
    """

    fields: dict[str, str]
    amount_field: str
    start_field: str
    end_field: str
    date_field: str = "date"
    label_field: str | None = None  # canonical id used to detect subtotal/total/group rows
    sheet_name: str | None = None  # None -> the active/first sheet
    header_row: int | None = None  # 1-based; None -> auto-detect by declared headers
    metadata_rows: tuple[int, int] | None = None  # 1-based inclusive top-block range
    date_format: str | None = None  # e.g. "%d.%m.%Y" for German plans
    total_markers: tuple[str, ...] = ("total",)  # label substrings -> subtotal/total skip


# ---------------------------------------------------------------------------
# Value coercion (dates / amounts).
# ---------------------------------------------------------------------------


def _parse_date(value: Any, *, date_format: str | None) -> date | None:
    """Coerce a cell to a date, or None if empty/unparseable.

    Accepts native datetime/date (openpyxl returns datetime for date cells), an
    optional declared format (e.g. %d.%m.%Y), then a few common fallbacks
    (ISO, German dotted). Returns None on empty so the caller decides noise.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    formats = [date_format] if date_format else []
    formats += ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y%m%d"]
    for fmt in formats:
        if not fmt:
            continue
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_amount(value: Any) -> Decimal | None:
    """Coerce an amount cell to a cent-quantised Decimal, or None if empty.

    Handles native numbers (incl. high-precision floats like the AXA plan's
    ``675622.4950015477``) and FR/EN-formatted strings ("1 234,56" / "1,234.56"
    with regular, NBSP, or thin-NBSP thousands separators). A non-empty but
    non-numeric cell returns None (the caller treats it as "no numeric amount").
    Quantised to the cent (ROUND_HALF_UP) so the sum-preserving invariant holds
    against a cent-exact line total.
    """
    if value is None or isinstance(value, bool):  # bool is an int subclass; never money
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(CENT, rounding=ROUND_HALF_UP)
        except InvalidOperation:
            return None
    text = str(value).strip()
    if not text:
        return None
    cleaned = "".join(ch for ch in text if not ch.isspace())
    for token in ("\N{EURO SIGN}", "EUR", "eur", "$"):
        cleaned = cleaned.replace(token, "")
    if "," in cleaned and "." in cleaned:
        # The LAST separator is the decimal point.
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")  # FR 1.234,56
        else:
            cleaned = cleaned.replace(",", "")  # EN 1,234.56
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        parsed = Decimal(cleaned)
    except InvalidOperation:
        return None
    if not parsed.is_finite():
        return None
    return parsed.quantize(CENT, rounding=ROUND_HALF_UP)


# ---------------------------------------------------------------------------
# Workbook / sheet helpers.
# ---------------------------------------------------------------------------


def _content_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _load_sheet(data: bytes, sheet_name: str | None) -> Any:
    """Load one worksheet from xlsx bytes (read_only=False so merges resolve)."""
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError as exc:  # pragma: no cover
        raise ReshapeProducerError(
            "openpyxl_unavailable", "openpyxl is required for Excel reshape."
        ) from exc
    try:
        # read_only=False so merged_cells ranges are available; data_only=True so
        # formula cells yield their last computed value, not the formula string.
        wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ReshapeProducerError(
            "unreadable_workbook",
            "The bytes are not a valid Excel (.xlsx) workbook.",
            repair={"export_values_only_xlsx": True},
        ) from exc
    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            raise ReshapeProducerError(
                "sheet_not_found",
                f"Sheet '{sheet_name}' not found.",
                repair={"pick_available_sheet": list(wb.sheetnames)},
            )
        return wb[sheet_name]
    ws = wb.active
    if ws is None:  # pragma: no cover - a loaded workbook always has an active sheet
        raise ReshapeProducerError("no_active_sheet", "Workbook has no active sheet.")
    return ws


def _detect_header_row(ws: Any, spec: ReshapeSpec) -> int:
    """Return the 1-based header row, explicit or auto-detected by declared headers.

    Auto-detection (spec.header_row is None): the first row (within
    HEADER_SCAN_LIMIT) on which EVERY declared column key resolves via the shared
    header resolver. Raises when no such row exists (the spec cannot bind).
    """
    if spec.header_row is not None:
        return spec.header_row
    keys = list(spec.fields.values())
    max_row = ws.max_row or 0
    scan_to = min(max_row, HEADER_SCAN_LIMIT)
    for candidate in range(1, scan_to + 1):
        if all(resolve_column_index(ws, candidate, key) is not None for key in keys):
            return candidate
    raise ReshapeProducerError(
        "no_header_row",
        "Could not locate a header row where every declared column is present. "
        "Declare an explicit header_row or check the column names.",
        repair={"declare_header_row": True},
    )


def _extract_metadata(
    ws: Any, merges: dict, rows: tuple[int, int]
) -> dict[str, Any]:
    """Extract the top ``label: value`` block over a 1-based inclusive row range.

    Each non-empty row contributes {first non-empty cell as label -> next
    non-empty cell as value}. Cells are merge-resolved. Rows with a single value
    (or none) are skipped. Interpreting the values (dates, budgets) is deferred.
    """
    start, end = rows
    out: dict[str, Any] = {}
    max_col = ws.max_column or 0
    for r in range(start, end + 1):
        values: list[str] = []
        for c in range(1, max_col + 1):
            raw, _ = cell_resolved(ws, merges, r, c)
            if raw is None:
                continue
            text = str(raw).strip()
            if text:
                values.append(text)
            if len(values) >= 2:
                break
        if len(values) >= 2:
            label = values[0].rstrip(":").strip()
            if label and label not in out:
                out[label] = values[1]
    return out


# ---------------------------------------------------------------------------
# Producer entrypoint.
# ---------------------------------------------------------------------------


def reshape_workbook_to_daily(
    data: bytes,
    spec: ReshapeSpec,
    *,
    sample_only: bool = False,
    sample_line_limit: int = 100,
) -> ParseResult:
    """Reshape one sheet's lines into DAILY canonical rows -> ``ParseResult``.

    Steps (all via the shared ``core.reshape`` engine):
      1. load the sheet, extract the declared top metadata block;
      2. detect the header row (explicit or auto by declared headers);
      3. resolve the declared columns + every merged cell;
      4. walk data rows: skip subtotal/total (label markers) and group/blank rows
         as honest ``RejectedRow`` evidence -- never a silent drop;
      5. explode each valid line's [start, end] range into daily rows whose
         ``amount_field`` values sum back to the line total EXACTLY (cent-exact).

    ``sample_only`` bounds the number of SOURCE lines processed (preview/gate).
    Raises ``ReshapeProducerError`` on structural failure (unreadable, no header,
    no data, missing declared columns, or a broken sum invariant).
    """
    if not data:
        raise ReshapeProducerError("empty_file", "The uploaded file is empty.")

    content_hash = _content_hash(data)
    ws = _load_sheet(data, spec.sheet_name)

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row > MAX_SHEET_ROWS or max_col > MAX_SHEET_COLUMNS:
        raise ReshapeProducerError(
            "sheet_too_large",
            f"Sheet is too large ({max_row}x{max_col}); "
            f"maximum {MAX_SHEET_ROWS}x{MAX_SHEET_COLUMNS}.",
        )

    merges = resolve_merges(ws)
    metadata = (
        _extract_metadata(ws, merges, spec.metadata_rows) if spec.metadata_rows else {}
    )

    header_row = _detect_header_row(ws, spec)

    # Resolve every declared field to a 1-based column index (once).
    field_col: dict[str, int] = {}
    for canonical_id, key in spec.fields.items():
        idx = resolve_column_index(ws, header_row, key)
        if idx is None:
            raise ReshapeProducerError(
                "column_not_found",
                f"Declared column '{key}' (field '{canonical_id}') not found on "
                f"header row {header_row}.",
                repair={"check_column_name_or_letter": key},
            )
        field_col[canonical_id] = idx

    for required in (spec.amount_field, spec.start_field, spec.end_field):
        if required not in field_col:
            raise ReshapeProducerError(
                "missing_required_column",
                f"The spec must map the '{required}' column.",
            )

    # Descriptive fields carried verbatim onto each daily row (everything mapped
    # except the two date-range columns, which collapse into the daily grain).
    carried_fields = [
        cid for cid in spec.fields if cid not in (spec.start_field, spec.end_field)
    ]
    # Label default: the first DESCRIPTIVE carried field (never the amount column,
    # so subtotal/total detection reads a name, not a number).
    _descriptive = [cid for cid in carried_fields if cid != spec.amount_field]
    label_id = spec.label_field or (_descriptive[0] if _descriptive else None)
    markers = tuple(m.strip().casefold() for m in spec.total_markers if m.strip())

    rows_out: list[dict[str, Any]] = []
    rejected: list[RejectedRow] = []
    source_lines_seen = 0

    first_data_row = header_row + 1
    last_row = ws.max_row or header_row
    for r in range(first_data_row, last_row + 1):
        # A fully-empty row is skipped silently (spacer), not counted as a line.
        if _row_is_blank(ws, merges, r, max_col):
            continue
        if sample_only and source_lines_seen >= sample_line_limit:
            break
        source_lines_seen += 1

        label = _read_text(ws, merges, field_col, label_id, r) if label_id else ""

        # Subtotal / grand-total row -> skip (kept as evidence, may validate later).
        if label and any(m in label.casefold() for m in markers):
            rejected.append(
                RejectedRow(
                    row_number=r, field_name=label_id or "",
                    rule=RULE_SUBTOTAL_OR_TOTAL,
                    reason=f"subtotal/total row skipped for landing: '{label}'",
                    rejected_value=_truncate(label),
                )
            )
            continue

        raw_amount, _ = cell_resolved(ws, merges, r, field_col[spec.amount_field])
        amount = _parse_amount(raw_amount)
        start = _parse_date(
            cell_resolved(ws, merges, r, field_col[spec.start_field])[0],
            date_format=spec.date_format,
        )
        end = _parse_date(
            cell_resolved(ws, merges, r, field_col[spec.end_field])[0],
            date_format=spec.date_format,
        )

        # Group-label / blank-ish row: no amount AND no dates -> noise.
        if amount is None and start is None and end is None:
            rejected.append(
                RejectedRow(
                    row_number=r, field_name=label_id or "",
                    rule=RULE_GROUP_OR_NOISE,
                    reason="group-label or empty row (no amount, no dates)",
                    rejected_value=_truncate(label),
                )
            )
            continue

        # An amount with no valid date range cannot be placed at the daily grain.
        if amount is not None and (start is None or end is None):
            rejected.append(
                RejectedRow(
                    row_number=r, field_name=spec.start_field,
                    rule=RULE_MISSING_DATES,
                    reason="line has an amount but no valid start/end date range",
                    rejected_value=_truncate(label),
                )
            )
            continue

        # Dates present but no numeric amount: a line we cannot spend-spread.
        if amount is None:
            rejected.append(
                RejectedRow(
                    row_number=r, field_name=spec.amount_field,
                    rule=RULE_BAD_AMOUNT,
                    reason="line has a date range but no numeric amount",
                    rejected_value=_truncate(str(raw_amount)),
                )
            )
            continue

        if start > end:  # type: ignore[operator]
            rejected.append(
                RejectedRow(
                    row_number=r, field_name=spec.start_field,
                    rule=RULE_INVERTED_DATES,
                    reason=f"start date {start} is after end date {end}",
                    rejected_value=_truncate(f"{start}..{end}"),
                )
            )
            continue

        # Carry the descriptive fields once (the same for every day of this line).
        base: dict[str, Any] = {}
        for cid in carried_fields:
            if cid == spec.amount_field:
                continue  # amount is per-day, filled below
            base[cid] = _read_text(ws, merges, field_col, cid, r) or None

        allocations = compute_spread(amount, start, end)  # cent-exact, sum-preserving
        total = sum((amt for _d, amt in allocations), Decimal("0"))
        if total != amount:  # defensive: compute_spread guarantees this
            raise ReshapeProducerError(
                "spread_invariant_broken",
                f"Row {r}: daily spread {total} != line total {amount}.",
            )
        for day, amt in allocations:
            rows_out.append(
                {**base, spec.date_field: day.isoformat(), spec.amount_field: format(amt, "f")}
            )

    if not rows_out and not rejected:
        raise ReshapeProducerError(
            "no_data_rows",
            f"No data rows found below header row {header_row}.",
        )

    columns = _synthesise_columns(spec, carried_fields)
    return ParseResult(
        rows=rows_out,
        rejected=rejected,
        columns=columns,
        encoding="xlsx",
        delimiter=None,
        sheet_name=ws.title,
        detected_row_count=source_lines_seen,
        content_hash=content_hash,
        metadata=metadata,
    )


# ---------------------------------------------------------------------------
# Small helpers.
# ---------------------------------------------------------------------------


def _row_is_blank(ws: Any, merges: dict, row: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        raw, _ = cell_resolved(ws, merges, row, c)
        if raw is not None and str(raw).strip() != "":
            return False
    return True


def _read_text(ws: Any, merges: dict, field_col: dict, canonical_id: str | None, row: int) -> str:
    if canonical_id is None or canonical_id not in field_col:
        return ""
    raw, _ = cell_resolved(ws, merges, row, field_col[canonical_id])
    return str(raw).strip() if raw is not None else ""


def _truncate(value: Any, max_len: int = 500) -> str:
    s = str(value) if value is not None else ""
    return s[:max_len] + "…" if len(s) > max_len else s


def _synthesise_columns(spec: ReshapeSpec, carried_fields: list[str]) -> list[ColumnSpec]:
    """Describe the EMITTED canonical fields (date grain + carried + amount)."""
    specs: list[ColumnSpec] = []
    idx = 0
    specs.append(ColumnSpec(name=spec.date_field, index=idx, detected_type="date"))
    for cid in carried_fields:
        idx += 1
        det = "decimal" if cid == spec.amount_field else "text"
        specs.append(ColumnSpec(name=cid, index=idx, detected_type=det))
    return specs


# ---------------------------------------------------------------------------
# Unified producer interface (Story 22.12, AD-1): one signature for both kinds,
# invoked at csv_excel_import.run_import's parse step. Returns the existing
# ParseResult with rows keyed by mdm_canonical_fields ids; everything downstream
# (managed_feed ledger + landing) is reused unchanged.
# ---------------------------------------------------------------------------


def _reshape_spec_from(reshape: dict[str, Any]) -> ReshapeSpec:
    """Build a ReshapeSpec from a template contract's ``reshape`` sub-document."""
    try:
        return ReshapeSpec(
            fields=dict(reshape["fields"]),
            amount_field=reshape["amount_field"],
            start_field=reshape["start_field"],
            end_field=reshape["end_field"],
            date_field=reshape.get("date_field", "date"),
            label_field=reshape.get("label_field"),
            sheet_name=reshape.get("sheet_name"),
            header_row=reshape.get("header_row"),
            metadata_rows=(
                tuple(reshape["metadata_rows"]) if reshape.get("metadata_rows") else None
            ),
            date_format=reshape.get("date_format"),
            total_markers=tuple(reshape.get("total_markers", ("total",))),
        )
    except (KeyError, TypeError) as exc:
        raise ReshapeProducerError(
            "invalid_reshape_spec",
            "the template's 'reshape' contract is missing required keys "
            "(fields/amount_field/start_field/end_field).",
        ) from exc


def _remap_to_canonical(parsed: ParseResult, mapping: dict[str, str]) -> ParseResult:
    """Rename a tabular ParseResult's columns to canonical field ids via ``mapping``.

    ``mapping`` is {source_column -> canonical_field_id}. Only mapped source
    columns are carried (extra columns are ignored -- never a silent drop of a
    REQUIRED field, which the 22.15 gate enforces separately). Rejected rows and
    the content hash pass through unchanged.
    """
    targets = sorted(mapping.items(), key=lambda kv: kv[1])
    new_rows = [{cid: row.get(src) for src, cid in targets} for row in parsed.rows]
    new_cols = [
        ColumnSpec(name=cid, index=i, detected_type="text")
        for i, (_src, cid) in enumerate(targets)
    ]
    return ParseResult(
        rows=new_rows,
        rejected=parsed.rejected,
        columns=new_cols,
        encoding=parsed.encoding,
        delimiter=parsed.delimiter,
        sheet_name=parsed.sheet_name,
        detected_row_count=parsed.detected_row_count,
        content_hash=parsed.content_hash,
        metadata=parsed.metadata,
    )


def produce(
    data: bytes,
    template: dict[str, Any],
    mapping: dict[str, str] | None = None,
    *,
    sample_only: bool = False,
) -> ParseResult:
    """Canonical-rows producer for a catalog template (AD-1 / AD-2).

    ONE signature for both template kinds. ``template`` is a file-source template
    row (or bare contract dict); ``mapping`` is {source_column -> canonical_field
    id} for the simple tabular path. Dispatch:

      * a contract carrying a ``reshape`` sub-document -> the media-plan-shaped
        Excel path (Story 22.10): reshape to daily canonical rows already keyed by
        canonical ids;
      * otherwise -> the simple tabular path: parse the CSV/Excel (Story 12.9) and
        rename columns to canonical ids via ``mapping``.

    Returns the existing ``ParseResult`` so ``run_import`` lands it through the
    unchanged ledger/landing chain -- no second parser, ledger, or landing path.
    """
    contract = template.get("contract") if "contract" in template else template
    if not isinstance(contract, dict):
        raise ReshapeProducerError("invalid_template", "template has no contract object.")

    reshape = contract.get("reshape")
    if reshape:
        spec = _reshape_spec_from(reshape)
        return reshape_workbook_to_daily(data, spec, sample_only=sample_only)

    # Simple tabular path: reuse the 12.9 parsers, then remap to canonical ids.
    from core.csv_excel_import import (  # noqa: PLC0415
        FORMAT_CSV,
        SUPPORTED_FORMATS,
        UnsupportedFileType,
        detect_format,
        parse_csv,
        parse_excel,
    )

    if not mapping:
        raise ReshapeProducerError(
            "mapping_required",
            "a tabular catalog template requires a source->canonical field mapping.",
        )
    header_row = int(contract.get("header_row", 1))
    date_format = contract.get("date_format")
    # Prefer the template's declared format; else detect from the bytes, falling
    # back to CSV (a header-less CSV has no magic bytes and no filename hint here).
    fmt = contract.get("format")
    if fmt not in SUPPORTED_FORMATS:
        try:
            fmt = detect_format(None, data)
        except UnsupportedFileType:
            fmt = FORMAT_CSV
    if fmt == FORMAT_CSV:
        parsed = parse_csv(data, header_row=header_row, date_format=date_format)
    else:
        parsed = parse_excel(
            data, sheet_name=contract.get("sheet_name"), header_row=header_row,
            date_format=date_format,
        )
    return _remap_to_canonical(parsed, mapping)


# ---------------------------------------------------------------------------
# Matrix placement (Story 22.14, AD-6): stamp every landed row with its class
# and any declared-discriminator dimension; refuse a template with no class.
# ---------------------------------------------------------------------------


def _resolve_discriminator(
    discriminator: dict[str, Any], *, filename: str | None, metadata: dict[str, Any]
) -> tuple[str | None, Any]:
    """Resolve a declared discriminator to (dimension_id, value).

    ``source`` = 'cell'     -> value = metadata[<cell label>] (from the top block);
                 'filename'  -> value = a named 'value' group (or group 1 / whole
                                match) of <pattern> applied to the filename.
    An unresolved discriminator yields (dimension_id, None) -- honest, never a
    fabricated value.
    """
    dim = discriminator.get("dimension")
    src = discriminator.get("source")
    value: Any = None
    if src == "cell":
        value = metadata.get(discriminator.get("cell"))
    elif src == "filename" and filename and discriminator.get("pattern"):
        match = re.search(discriminator["pattern"], filename)
        if match:
            named = match.groupdict().get("value")
            value = named if named is not None else (
                match.group(1) if match.groups() else match.group(0)
            )
    return dim, value


def evaluate_variant_discriminator(
    template: dict[str, Any],
    *,
    filename: str | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Resolve a declared variant discriminator and flag it if undetectable (Story 22.18).

    Several country/market variants (FR, DE, ...) land in ONE datastream: a
    declared discriminator (a named metadata cell or a filename pattern) is
    promoted to a market/country dimension. When the discriminator is DECLARED but
    its value cannot be detected on this file, the import is FLAGGED for human
    confirmation rather than landing a mislabeled / null dimension.

    Returns {dimension, value, detected, flagged, reason}. A template that declares
    no discriminator is not a variant datastream -> nothing to flag.
    """
    contract = _template_contract(template)
    discriminator = contract.get("discriminator")
    if not isinstance(discriminator, dict):
        return {"dimension": None, "value": None, "detected": False,
                "flagged": False, "reason": None}

    dim, value = _resolve_discriminator(
        discriminator, filename=filename, metadata=metadata or {}
    )
    detected = value is not None and str(value).strip() != ""
    return {
        "dimension": dim,
        "value": value if detected else None,
        "detected": detected,
        "flagged": not detected,  # declared but undetectable -> confirm, don't mislabel
        "reason": None if detected else "undetectable_discriminator",
    }


def stamp_placement(
    result: ParseResult,
    template: dict[str, Any],
    *,
    filename: str | None = None,
) -> ParseResult:
    """Stamp every row with its matrix class + declared-discriminator dimension.

    AD-6: an import whose template declares NO class is refused BEFORE landing
    (a ReshapeProducerError, not a silent pass). The class (planned/actual/
    extrapolated) is stamped on ``_placement_class``; an optional declared
    discriminator (a named cell value or a filename pattern) is promoted to its
    declared canonical dimension on every row.

    Mutates and returns ``result`` (rows carry the placement so downstream can
    reconcile plan-vs-actual on the same coordinates).
    """
    contract = template.get("contract") if "contract" in template else template
    if not isinstance(contract, dict):
        raise ReshapeProducerError("invalid_template", "template has no contract object.")

    placement_class = contract.get("class")
    if placement_class not in PLACEMENT_CLASSES:
        raise ReshapeProducerError(
            "no_placement_class",
            "the template declares no matrix class (planned/actual/extrapolated); "
            "the import is refused before landing (AD-6).",
        )

    discriminator = contract.get("discriminator")
    disc_dim, disc_val = (None, None)
    if isinstance(discriminator, dict):
        disc_dim, disc_val = _resolve_discriminator(
            discriminator, filename=filename, metadata=result.metadata or {}
        )

    for row in result.rows:
        row[PLACEMENT_CLASS_KEY] = placement_class
        if disc_dim:
            row[disc_dim] = disc_val
    return result


def landed_total(result: ParseResult, amount_field: str) -> Decimal:
    """Sum a canonical amount field across the landed rows (Decimal, exact).

    Used to prove the F-1 reconciliation invariant (landed total == file total)
    without re-deriving amounts through a different path.
    """
    total = Decimal("0.00")
    for row in result.rows:
        raw = row.get(amount_field)
        if raw is None:
            continue
        total += Decimal(str(raw))
    return total


# ---------------------------------------------------------------------------
# Lock, replay & drift (Story 22.16, AD-8): a locked template replays
# byte-identically by upload and by email; source drift re-enters the gate.
# ---------------------------------------------------------------------------

DRIFT_OK = "ok"  # required source columns all present -> deterministic replay
DRIFT_NEEDS_REVALIDATION = "needs_revalidation"  # a required source disappeared -> gate


class TemplateNotLocked(ReshapeProducerError):
    """Replay was requested on a template that is not locked (no content_hash)."""

    def __init__(self) -> None:
        super().__init__(
            "template_not_locked",
            "replay requires a LOCKED, content-hashed template (Story 22.11).",
        )


def _template_contract(template: dict[str, Any]) -> dict[str, Any]:
    contract = template.get("contract") if "contract" in template else template
    if not isinstance(contract, dict):
        raise ReshapeProducerError("invalid_template", "template has no contract object.")
    return contract


def canonical_rows_signature(result: ParseResult) -> str:
    """A stable SHA-256 over the landed rows -- the 'byte-identical' witness.

    Two replays of the SAME bytes through the SAME locked template (whatever the
    ingress -- upload or email) yield the SAME signature (AD-8 idempotent replay).
    """
    canonical = json.dumps(result.rows, sort_keys=True, separators=(",", ":"),
                           ensure_ascii=False, default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def read_source_columns(data: bytes, template: dict[str, Any]) -> list[str]:
    """Return the source header column names of ``data`` under this template.

    Used to detect source drift (added / reordered / disappeared columns). Works
    for both the reshape (media-plan) path and the simple tabular path.
    """
    contract = _template_contract(template)
    reshape = contract.get("reshape")
    if reshape:
        ws = _load_sheet(data, reshape.get("sheet_name"))
        header_row = reshape.get("header_row")
        if header_row is None:
            header_row = _detect_header_row(ws, _reshape_spec_from(reshape))
        merges = resolve_merges(ws)
        cols: list[str] = []
        for c in range(1, (ws.max_column or 0) + 1):
            raw, _ = cell_resolved(ws, merges, header_row, c)
            if raw is not None and str(raw).strip():
                cols.append(str(raw).strip())
        return cols

    from core.csv_excel_import import (  # noqa: PLC0415
        FORMAT_CSV,
        SUPPORTED_FORMATS,
        UnsupportedFileType,
        detect_format,
        parse_csv,
        parse_excel,
    )

    header_row = int(contract.get("header_row", 1))
    fmt = contract.get("format")
    if fmt not in SUPPORTED_FORMATS:
        try:
            fmt = detect_format(None, data)
        except UnsupportedFileType:
            fmt = FORMAT_CSV
    parsed = (
        parse_csv(data, header_row=header_row) if fmt == FORMAT_CSV
        else parse_excel(data, sheet_name=contract.get("sheet_name"), header_row=header_row)
    )
    return [c.name for c in parsed.columns]


def detect_source_drift(
    template: dict[str, Any], mapping: dict[str, str], current_columns: list[str]
) -> dict[str, Any]:
    """Classify a file's source columns against the locked template + mapping (AD-8).

    Mapping is by column NAME, so an ADDED or REORDERED column is harmless -- the
    required fields still resolve and land. A DISAPPEARED required source column
    (a column feeding a required canonical field is gone) returns
    ``needs_revalidation``: the file must re-enter the AD-7 gate for human
    re-confirmation, never a silent re-map.

    Returns {status, missing_required_sources, added_columns}.
    """
    contract = _template_contract(template)
    required = set(contract.get("required_fields") or [])
    current = set(current_columns)

    missing_required_sources = sorted(
        source_col
        for source_col, canonical_id in (mapping or {}).items()
        if canonical_id in required and source_col not in current
    )
    added_columns = sorted(current - set(mapping or {}))
    status = DRIFT_NEEDS_REVALIDATION if missing_required_sources else DRIFT_OK
    return {
        "status": status,
        "missing_required_sources": missing_required_sources,
        "added_columns": added_columns,
    }


def replay_locked_template(
    data: bytes,
    template: dict[str, Any],
    mapping: dict[str, str] | None = None,
    *,
    filename: str | None = None,
) -> ParseResult:
    """Deterministically replay a LOCKED template over ``data`` (AD-8).

    Ingress parity: upload and email are only transports of the same bytes, so the
    same (locked template + mapping + bytes) yields byte-identical canonical rows
    (see ``canonical_rows_signature``). Requires the template to be locked
    (content-hashed, Story 22.11) -- an unlocked template is refused.
    """
    if not template.get("content_hash"):
        raise TemplateNotLocked()
    return stamp_placement(produce(data, template, mapping), template, filename=filename)
